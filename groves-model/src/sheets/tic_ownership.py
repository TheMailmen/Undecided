# src/sheets/tic_ownership.py — TIC Ownership & Returns sheet builder
"""
Builds the 'TIC Ownership' tab: co-owner table with return metrics,
property summary, and T12 financial highlights.

Exact layout (27 rows × 8 cols) required by test_model.py:
  Row  1–2  : title / subtitle bars
  Row  3    : column headers
  Rows 4–6  : TIC entities (cols E–H must be formulas)
  Row  7    : Totals row  (col C = SUM of ownership % ≈ 1.0)
  Row  8    : spacer
  Row  9    : PROPERTY SUMMARY section
  Rows 10–14: property details
  Row 15    : spacer
  Row 16    : T12 FINANCIALS section
  Rows 17–23: T12 metrics  (row 21 col C = NOI, cross-ref'd from T12_PL col 15)
  Row 24    : spacer
  Row 25    : VALUATION section
  Rows 26–27: valuation metrics

Column map:
  A (1): Entity / Label
  B (2): Notes / blank
  C (3): Ownership % [entities] | Value [property & T12 rows]
  D (4): Equity Invested [entities]
  E (5): Annual NOI Allocation  = $C$21 × C  [formula ✓]
  F (6): Annual CFADS Allocation = $C$22 × C  [formula ✓]
  G (7): Cash-on-Cash Return    = F / D       [formula ✓]
  H (8): Estimated Equity Value = $C$26 × C   [formula ✓]
"""
from openpyxl.utils import get_column_letter

from design import (
    apply_title, apply_hdr, apply_section,
    hide_beyond, input_cell,
    F_BOLD, F_BODY, F_MUTED, F_LABEL, F_GREEN,
    P_ALT, P_WHITE, P_GREEN_LT, P_SUBTOTAL,
    B_SUB,
    NF_DOLLAR, NF_PCT, NF_DEC2,
    A_CENTER, A_LEFT, A_RIGHT,
)

MAX_COL = 8

# T12_PL column references (col O = 15 = T12 total)
_T12_SHEET    = 'T12_PL'
_T12_TOTAL_COL = get_column_letter(15)   # O

# Row numbers (1-indexed) — must match the layout exactly
_ROW_TITLE    = 1
_ROW_SUB      = 2
_ROW_HDR      = 3
_ROW_ENT_1    = 4
_ROW_ENT_2    = 5
_ROW_ENT_3    = 6
_ROW_TOTAL    = 7
_ROW_NOI      = 21   # T12 NOI  — test checks ws.cell(21, 3)
_ROW_CFADS    = 22   # T12 CFADS
_ROW_BOV      = 26   # BOV midpoint


def _t12_lookup(account):
    """Excel INDEX/MATCH formula pulling T12 total from T12_PL col O."""
    return (
        f'=INDEX({_T12_SHEET}!${_T12_TOTAL_COL}:${_T12_TOTAL_COL},'
        f'MATCH("{account}",{_T12_SHEET}!$B:$B,0))'
    )


def build(wb, cfg):
    ws = wb.create_sheet('TIC Ownership')

    prop   = cfg['property']
    tic    = cfg['tic']
    equity = cfg['total_equity']
    val    = cfg['valuation']
    loan   = cfg['loan']

    # ordered list of (name, pct, invested_equity)
    entities = [(name, v['pct'], v['equity']) for name, v in tic.items()]

    # ── Title / subtitle ──────────────────────────────────────────────
    apply_title(
        ws, MAX_COL,
        f"{prop['name']}  |  TIC Ownership & Returns",
        f"Tenants in Common  |  {len(entities)} Co-owners  |  "
        f"${equity:,.0f} Total Equity",
    )

    # ── Column headers (row 3) ────────────────────────────────────────
    apply_hdr(ws, _ROW_HDR, MAX_COL)
    for col, label in [
        (1, 'Entity'),
        (2, ''),
        (3, 'Ownership %'),
        (4, 'Equity Invested'),
        (5, 'Annual NOI Share'),
        (6, 'Annual CFADS Share'),
        (7, 'Cash-on-Cash'),
        (8, 'Est. Equity Value'),
    ]:
        cell = ws.cell(row=_ROW_HDR, column=col)
        cell.value = label
        cell.alignment = A_CENTER

    # ── TIC entity rows (4–6) ─────────────────────────────────────────
    for i, (name, pct, invested) in enumerate(entities):
        r   = _ROW_ENT_1 + i
        fill = P_ALT if r % 2 == 0 else P_WHITE
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill

        ws.cell(row=r, column=1).value = name
        ws.cell(row=r, column=1).font  = F_BOLD

        # C: ownership % (static input)
        pct_cell = input_cell(ws, r, 3, value=pct, nf=NF_PCT)

        # D: equity invested (static input)
        input_cell(ws, r, 4, value=invested, nf=NF_DOLLAR)

        # E: Annual NOI share (formula — col 5, refs row 21 col C)
        e_cell = ws.cell(row=r, column=5)
        e_cell.value = f'=$C${_ROW_NOI}*C{r}'
        e_cell.number_format = NF_DOLLAR

        # F: Annual CFADS share (formula — col 6, refs row 22 col C)
        f_cell = ws.cell(row=r, column=6)
        f_cell.value = f'=$C${_ROW_CFADS}*C{r}'
        f_cell.number_format = NF_DOLLAR

        # G: Cash-on-Cash return (formula — col 7)
        g_cell = ws.cell(row=r, column=7)
        g_cell.value = f'=IF(D{r}>0,F{r}/D{r},0)'
        g_cell.number_format = NF_PCT

        # H: Estimated equity value (formula — col 8, refs row 26 col C)
        h_cell = ws.cell(row=r, column=8)
        h_cell.value = f'=$C${_ROW_BOV}*C{r}'
        h_cell.number_format = NF_DOLLAR

    # ── Totals row (7) ────────────────────────────────────────────────
    r = _ROW_TOTAL
    for c in range(1, MAX_COL + 1):
        ws.cell(row=r, column=c).fill = P_SUBTOTAL
        ws.cell(row=r, column=c).border = B_SUB

    ws.cell(row=r, column=1).value = 'TOTAL'
    ws.cell(row=r, column=1).font  = F_BOLD

    for col, nf in [(3, NF_PCT), (4, NF_DOLLAR), (5, NF_DOLLAR),
                    (6, NF_DOLLAR), (8, NF_DOLLAR)]:
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f'=SUM({cl}{_ROW_ENT_1}:{cl}{_ROW_ENT_3})'
        cell.number_format = nf
        cell.font = F_BOLD

    # G7: aggregate CoC
    ws.cell(row=r, column=7).value = f'=IF(D{r}>0,F{r}/D{r},0)'
    ws.cell(row=r, column=7).number_format = NF_PCT
    ws.cell(row=r, column=7).font = F_BOLD

    # ── Spacer row 8 ──────────────────────────────────────────────────
    # (empty — no fill needed)

    # ── PROPERTY SUMMARY section (row 9) ─────────────────────────────
    apply_section(ws, 9, MAX_COL, accent=True)
    ws.cell(row=9, column=1).value = 'PROPERTY SUMMARY'

    def _prop_row(r_num, label, value, nf=None):
        fill = P_ALT if r_num % 2 == 0 else P_WHITE
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r_num, column=c).fill = fill
        ws.cell(row=r_num, column=1).value = label
        ws.cell(row=r_num, column=1).font  = F_LABEL
        cell_c = ws.cell(row=r_num, column=3)
        cell_c.value = value
        cell_c.font  = F_BODY
        if nf:
            cell_c.number_format = nf

    _prop_row(10, 'Property',       prop['name'])
    _prop_row(11, 'Address',        prop['address'])
    _prop_row(12, 'Units',          prop['units'])
    _prop_row(13, 'Purchase Price', prop['purchase_price'], NF_DOLLAR)
    _prop_row(14, 'Loan Balance',   loan['est_current_balance'], NF_DOLLAR)
    # row 15 = spacer

    # ── T12 FINANCIALS section (row 16) ──────────────────────────────
    apply_section(ws, 16, MAX_COL, accent=True)
    ws.cell(row=16, column=1).value = 'T12 FINANCIALS'

    t12_rows = [
        (17, 'Gross Potential Rent',             'Gross Potential Rent',            NF_DOLLAR),
        (18, 'Effective Gross Income',            'EFFECTIVE GROSS INCOME (EGI)',    NF_DOLLAR),
        (19, 'Total Operating Expenses',          'Total Operating Expenses',        NF_DOLLAR),
        (20, 'Total Debt Service',                'Total Debt Service',              NF_DOLLAR),
        (_ROW_NOI,   'NET OPERATING INCOME (NOI)',  'NET OPERATING INCOME (NOI)',    NF_DOLLAR),  # row 21
        (_ROW_CFADS, 'Cash Flow After Debt Svc',   'CASH FLOW AFTER DEBT SERVICE',  NF_DOLLAR),  # row 22
        (23, 'Net Cash Flow',                     'NET CASH FLOW',                  NF_DOLLAR),
    ]

    for r_num, label, account, nf in t12_rows:
        is_green = account in (
            'NET OPERATING INCOME (NOI)', 'CASH FLOW AFTER DEBT SERVICE', 'NET CASH FLOW'
        )
        fill = P_GREEN_LT if is_green else (P_ALT if r_num % 2 == 0 else P_WHITE)
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r_num, column=c).fill = fill

        ws.cell(row=r_num, column=1).value = label
        ws.cell(row=r_num, column=1).font  = F_GREEN if is_green else F_LABEL

        cell_c = ws.cell(row=r_num, column=3)
        cell_c.value = _t12_lookup(account)
        cell_c.number_format = nf
        cell_c.font = F_GREEN if is_green else F_BODY

    # row 24 = spacer

    # ── VALUATION section (row 25) ────────────────────────────────────
    apply_section(ws, 25, MAX_COL, accent=True)
    ws.cell(row=25, column=1).value = 'VALUATION'

    # Row 26: BOV midpoint (input — referenced by H formulas above)
    fill26 = P_ALT if 26 % 2 == 0 else P_WHITE
    for c in range(1, MAX_COL + 1):
        ws.cell(row=26, column=c).fill = fill26
    ws.cell(row=26, column=1).value = 'BOV Midpoint'
    ws.cell(row=26, column=1).font  = F_LABEL
    input_cell(ws, 26, 3, value=val['bov_mid'], nf=NF_DOLLAR)

    # Row 27: Implied cap rate
    fill27 = P_WHITE if 27 % 2 == 0 else P_ALT
    for c in range(1, MAX_COL + 1):
        ws.cell(row=27, column=c).fill = fill27
    ws.cell(row=27, column=1).value = 'Implied Cap Rate'
    ws.cell(row=27, column=1).font  = F_LABEL
    ws.cell(row=27, column=3).value = f'=IF(C{_ROW_BOV}>0,C{_ROW_NOI}/C{_ROW_BOV},0)'
    ws.cell(row=27, column=3).number_format = NF_PCT

    # ── Column widths ─────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 4
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 18

    # ── Hide beyond data boundary ─────────────────────────────────────
    hide_beyond(ws, 27, MAX_COL)

    return ws
