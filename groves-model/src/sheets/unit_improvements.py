# src/sheets/unit_improvements.py — Unit Improvements sheet builder
"""
Builds the 'Unit Improvements' tab: renovation tracking from
data/unit_improvements.csv.

CSV columns: Unit,Mix,Condition,OrigRent,CurrRent,MktRent,LTL,
             Stove,Fridge,Dishwasher,Sink,Microwave,Flooring,
             Cabinets,Countertops,Toilet,Vanity,Tub,TubSurround,Notes
"""
import csv, os

from design import (
    apply_title, apply_hdr, apply_section, apply_subtotal,
    hide_beyond,
    F_BOLD, F_BODY, F_LABEL, F_MUTED, F_GREEN,
    P_ALT, P_WHITE, P_SUBTOTAL, P_GREEN_LT,
    NF_DOLLAR, NF_PCT,
    A_CENTER,
)


def build(wb, cfg, ui_csv_path=None):
    import os
    if ui_csv_path is None:
        ui_csv_path = os.path.join(
            os.path.dirname(__file__), '..', '..', 'data', 'unit_improvements.csv'
        )

    ws = wb.create_sheet('Unit Improvements')
    prop = cfg['property']

    rows = []
    headers = []
    if os.path.exists(ui_csv_path):
        with open(ui_csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            for row in reader:
                rows.append(row)

    # Columns to display (subset of CSV columns)
    display_cols = ['Unit', 'Mix', 'Condition', 'OrigRent', 'CurrRent',
                    'MktRent', 'LTL', 'Flooring', 'Cabinets', 'Notes']
    display_cols = [c for c in display_cols if c in headers] if headers else display_cols
    MAX_COL = len(display_cols)

    apply_title(
        ws, MAX_COL,
        f"{prop['name']}  |  Unit Improvements",
        f"Renovation scope and rent impact by unit  |  {len(rows)} renovated units",
    )

    # ── Column headers (row 3) ────────────────────────────────────────
    apply_hdr(ws, 3, MAX_COL)
    for i, col in enumerate(display_cols, 1):
        ws.cell(row=3, column=i).value = col
        ws.cell(row=3, column=i).alignment = A_CENTER

    # ── Data rows ─────────────────────────────────────────────────────
    DATA_START = 4
    for idx, row in enumerate(rows):
        r    = DATA_START + idx
        fill = P_ALT if r % 2 == 0 else P_WHITE
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill

        for i, col in enumerate(display_cols, 1):
            val = row.get(col, '')
            cell = ws.cell(row=r, column=i)
            if col in ('OrigRent', 'CurrRent', 'MktRent', 'LTL'):
                try:
                    cell.value = float(val) if val else ''
                    cell.number_format = NF_DOLLAR
                except ValueError:
                    cell.value = val
                cell.font = F_BODY
            else:
                cell.value = val
                cell.font = F_MUTED if col == 'Unit' else F_BODY

    last_row = DATA_START + len(rows) - 1 if rows else DATA_START

    # ── Summary row ───────────────────────────────────────────────────
    if rows:
        sum_row = last_row + 1
        apply_subtotal(ws, sum_row, MAX_COL)
        ws.cell(row=sum_row, column=1).value = f'TOTAL RENOVATED: {len(rows)} units'
        ws.cell(row=sum_row, column=1).font  = F_BOLD
        # Avg rent uplift
        for col_name, col_idx in [('CurrRent', display_cols.index('CurrRent') + 1
                                    if 'CurrRent' in display_cols else None),
                                   ('MktRent',  display_cols.index('MktRent') + 1
                                    if 'MktRent' in display_cols else None)]:
            if col_idx:
                ltr = chr(64 + col_idx)
                ws.cell(row=sum_row, column=col_idx).value = (
                    f'=AVERAGE({ltr}{DATA_START}:{ltr}{last_row})'
                )
                ws.cell(row=sum_row, column=col_idx).number_format = NF_DOLLAR
        last_row = sum_row

    # ── Column widths ─────────────────────────────────────────────────
    widths = {'Unit': 8, 'Mix': 8, 'Condition': 14, 'OrigRent': 10,
              'CurrRent': 10, 'MktRent': 10, 'LTL': 8,
              'Flooring': 10, 'Cabinets': 10, 'Notes': 30}
    for i, col in enumerate(display_cols, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 12)

    hide_beyond(ws, last_row, MAX_COL)
    return ws
