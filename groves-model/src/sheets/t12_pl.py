# src/sheets/t12_pl.py — Trailing 12-Month P&L sheet builder
"""
Builds the 'T12_PL' tab: the most recent 12 months of P&L actuals
with a T12 total column at the end (column 15 = col O).

Column layout:
  A  (1)  - GL Code
  B  (2)  - Account Name
  C–N (3–14) - Monthly columns (Jan-25 through Dec-25)
  O  (15) - T12 Total (SUM of cols C–N)
"""
from openpyxl.utils import get_column_letter

from design import (
    apply_title, apply_hdr, apply_pl_formatting, hide_beyond,
    NF_DOLLAR, NF_PCT, NF_DEC2, NF_DATE,
    A_CENTER, F_BOLD,
)

_METRIC_PCT   = {'Cap Rate', 'Expense Ratio', 'Cash-on-Cash (CFADS)', 'Cash-on-Cash (Ann.)'}
_METRIC_RATIO = {'DSCR'}
_T12_MONTHS   = 12
_TOTAL_COL    = 15   # column O — T12 total; test validates NOI here


def build(wb, cfg):
    ws = wb.create_sheet('T12_PL')

    # ── Pull the 12 most recent months from qPL_Fact ─────────────────
    qpl_ws = wb['qPL_Fact']
    seen, all_months = set(), []
    for row in qpl_ws.iter_rows(min_row=2, values_only=True):
        m = row[0]
        if m and m not in seen:
            seen.add(m)
            all_months.append(m)
    all_months.sort()

    months = all_months[-_T12_MONTHS:]   # last 12
    n      = len(months)
    FIRST_MO_COL = 3                     # column C

    # last_col should equal _TOTAL_COL (15) when n == 12
    last_col = FIRST_MO_COL + n          # monthly cols + 1 total col

    # ── Title / subtitle ──────────────────────────────────────────────
    prop     = cfg['property']
    first_mo = months[0].strftime('%b %Y')  if hasattr(months[0], 'strftime') else str(months[0])
    last_mo  = months[-1].strftime('%b %Y') if hasattr(months[-1], 'strftime') else str(months[-1])
    apply_title(
        ws, last_col,
        f"{prop['name']}  |  Trailing 12-Month P&L",
        f"{first_mo} – {last_mo}  |  {prop['address']}",
    )

    # ── Column headers (row 3) ────────────────────────────────────────
    apply_hdr(ws, 3, last_col)
    ws.cell(row=3, column=1).value = 'GL'
    ws.cell(row=3, column=2).value = 'Account'

    for i, m in enumerate(months):
        cell = ws.cell(row=3, column=FIRST_MO_COL + i)
        cell.value = m
        cell.number_format = NF_DATE
        cell.alignment = A_CENTER

    total_hdr = ws.cell(row=3, column=last_col)
    total_hdr.value = 'T12 Total'
    total_hdr.alignment = A_CENTER

    # ── Write COA rows (starting row 4) ──────────────────────────────
    coa        = cfg['coa']
    data_start = 4
    r          = data_start

    first_mo_ltr  = get_column_letter(FIRST_MO_COL)
    last_mo_ltr   = get_column_letter(FIRST_MO_COL + n - 1)
    total_col_ltr = get_column_letter(last_col)

    for gl, acct, rtype in coa:
        ws.cell(row=r, column=1).value = gl
        ws.cell(row=r, column=2).value = acct

        if rtype == 'spacer' or acct is None:
            r += 1
            continue

        if acct in _METRIC_PCT:
            nf = NF_PCT
        elif acct in _METRIC_RATIO:
            nf = NF_DEC2
        else:
            nf = NF_DOLLAR

        # Monthly SUMIFS
        for i in range(n):
            col    = FIRST_MO_COL + i
            mo_ltr = get_column_letter(col)
            cell   = ws.cell(row=r, column=col)
            cell.value = (
                f'=SUMIFS(qPL_Fact!$D:$D,'
                f'qPL_Fact!$C:$C,$B{r},'
                f'qPL_Fact!$A:$A,{mo_ltr}$3)'
            )
            cell.number_format = nf

        # T12 total column
        total_cell = ws.cell(row=r, column=last_col)
        if rtype == 'metric':
            # metrics: last-month value (ratios don't sum)
            total_cell.value = f'={last_mo_ltr}{r}'
        else:
            total_cell.value = f'=SUM({first_mo_ltr}{r}:{last_mo_ltr}{r})'
        total_cell.number_format = nf

        r += 1

    last_row = r - 1

    # ── Apply section / subtotal / row formatting ─────────────────────
    apply_pl_formatting(ws, last_col, data_start=data_start, data_end=last_row)

    # ── Column widths ─────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 32
    for i in range(n):
        ws.column_dimensions[get_column_letter(FIRST_MO_COL + i)].width = 11
    ws.column_dimensions[total_col_ltr].width = 13

    # ── Freeze panes ─────────────────────────────────────────────────
    ws.freeze_panes = 'C4'

    # ── Hide beyond data ──────────────────────────────────────────────
    hide_beyond(ws, last_row, last_col)

    return ws
