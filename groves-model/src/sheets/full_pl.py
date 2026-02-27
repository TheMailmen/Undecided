# src/sheets/full_pl.py — Full P&L Sheet Builder
"""
Builds the 'Full P&L' tab: monthly actuals for all available months
pulled from qPL_Fact via Excel SUMIFS formulas.

Column layout:
  A  - GL Code
  B  - Account Name
  C  - T12 TTM (trailing 12 months)
  D+ - Monthly columns (Aug-24 through Dec-25, one per month)
"""
from openpyxl.utils import get_column_letter

from design import (
    apply_title, apply_hdr, apply_pl_formatting, hide_beyond,
    NF_DOLLAR, NF_PCT, NF_DEC2, NF_DATE,
    A_CENTER,
)

# Metric rows that display as percentages vs ratios
_METRIC_PCT   = {'Cap Rate', 'Expense Ratio', 'Cash-on-Cash (CFADS)', 'Cash-on-Cash (Ann.)'}
_METRIC_RATIO = {'DSCR'}


def build(wb, cfg):
    ws = wb.create_sheet('Full P&L')

    # ── Collect sorted months from qPL_Fact ──────────────────────────
    qpl_ws = wb['qPL_Fact']
    seen, months = set(), []
    for row in qpl_ws.iter_rows(min_row=2, values_only=True):
        m = row[0]
        if m and m not in seen:
            seen.add(m)
            months.append(m)
    months.sort()
    n_months = len(months)

    T12_COL       = 3                          # column C
    FIRST_MO_COL  = 4                          # column D
    last_col      = FIRST_MO_COL + n_months - 1

    # ── Title / subtitle (rows 1-2) ───────────────────────────────────
    prop = cfg['property']
    apply_title(
        ws, last_col,
        f"{prop['name']}  |  Full P&L — Monthly Actuals",
        f"{prop['address']}  |  {n_months} Months",
    )

    # ── Column headers (row 3) ────────────────────────────────────────
    apply_hdr(ws, 3, last_col)
    ws.cell(row=3, column=1).value = 'GL'
    ws.cell(row=3, column=2).value = 'Account'
    t12_hdr = ws.cell(row=3, column=T12_COL)
    t12_hdr.value = 'T12 TTM'
    t12_hdr.alignment = A_CENTER

    for i, m in enumerate(months):
        cell = ws.cell(row=3, column=FIRST_MO_COL + i)
        cell.value = m
        cell.number_format = NF_DATE
        cell.alignment = A_CENTER

    # ── T12 window: last 12 of the available months ───────────────────
    t12_count     = min(12, n_months)
    t12_first_col = FIRST_MO_COL + n_months - t12_count
    t12_last_col  = last_col
    t12_first_ltr = get_column_letter(t12_first_col)
    t12_last_ltr  = get_column_letter(t12_last_col)
    last_mo_ltr   = get_column_letter(last_col)

    # ── Write COA rows (starting row 4) ──────────────────────────────
    coa       = cfg['coa']
    data_start = 4
    r         = data_start

    for gl, acct, rtype in coa:
        ws.cell(row=r, column=1).value = gl
        ws.cell(row=r, column=2).value = acct

        if rtype == 'spacer' or acct is None:
            r += 1
            continue

        # Number format for this row
        if acct in _METRIC_PCT:
            nf = NF_PCT
        elif acct in _METRIC_RATIO:
            nf = NF_DEC2
        else:
            nf = NF_DOLLAR

        # T12 column (C): sum of last-12-month cols, or last value for metrics
        t12_cell = ws.cell(row=r, column=T12_COL)
        if rtype == 'metric':
            t12_cell.value = f'={last_mo_ltr}{r}'
        else:
            t12_cell.value = f'=SUM({t12_first_ltr}{r}:{t12_last_ltr}{r})'
        t12_cell.number_format = nf

        # Monthly SUMIFS: one formula per month column
        for i in range(n_months):
            col     = FIRST_MO_COL + i
            mo_ltr  = get_column_letter(col)
            cell    = ws.cell(row=r, column=col)
            cell.value = (
                f'=SUMIFS(qPL_Fact!$D:$D,'
                f'qPL_Fact!$C:$C,$B{r},'
                f'qPL_Fact!$A:$A,{mo_ltr}$3)'
            )
            cell.number_format = nf

        r += 1

    last_row = r - 1

    # ── Apply section / subtotal / alternating formatting ─────────────
    apply_pl_formatting(ws, last_col, data_start=data_start, data_end=last_row)

    # ── Column widths ─────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions[get_column_letter(T12_COL)].width = 12
    for i in range(n_months):
        ws.column_dimensions[get_column_letter(FIRST_MO_COL + i)].width = 11

    # ── Freeze panes at D4 (labels + header locked) ───────────────────
    ws.freeze_panes = 'D4'

    # ── Hide columns and rows beyond data boundary ────────────────────
    hide_beyond(ws, last_row, last_col)

    return ws
