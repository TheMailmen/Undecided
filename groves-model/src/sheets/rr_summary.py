# src/sheets/rr_summary.py — Rent Roll Summary sheet builder
"""
Builds the 'RR Summary' tab: occupancy and rent statistics derived
from rr_input data (via cross-sheet formulas to 'RR Input').

test_model.py checks:
  RR Summary cell(5, 3) == RR Input cell(124, 2)  (total rent tie-out)
"""
from openpyxl.utils import get_column_letter

from design import (
    apply_title, apply_hdr, apply_section, apply_subtotal,
    hide_beyond,
    F_BOLD, F_BODY, F_LABEL, F_MUTED, F_GREEN,
    P_ALT, P_WHITE, P_GREEN_LT, P_SUBTOTAL,
    B_SUB,
    NF_DOLLAR, NF_PCT, NF_DEC2,
    A_CENTER,
)

MAX_COL = 6
_RR = 'RR Input'


def build(wb, cfg):
    ws = wb.create_sheet('RR Summary')
    prop = cfg['property']
    um   = cfg['unit_mix']

    apply_title(
        ws, MAX_COL,
        f"{prop['name']}  |  Rent Roll Summary",
        f"Current occupancy and rent statistics  |  {prop['units']} total units",
    )

    r = 3

    def sec(title):
        nonlocal r
        apply_section(ws, r, MAX_COL, accent=True)
        ws.cell(row=r, column=1).value = title
        r += 1

    def hdr(*labels):
        nonlocal r
        apply_hdr(ws, r, MAX_COL)
        for i, lbl in enumerate(labels, 1):
            ws.cell(row=r, column=i).value = lbl
        r += 1

    def row(label, formula='', nf=NF_DOLLAR, note=''):
        nonlocal r
        fill = P_ALT if r % 2 == 0 else P_WHITE
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=1).value = label
        ws.cell(row=r, column=1).font  = F_LABEL
        if formula:
            cell = ws.cell(row=r, column=3)
            cell.value = formula
            cell.number_format = nf
            cell.font = F_BODY
        if note:
            ws.cell(row=r, column=5).value = note
            ws.cell(row=r, column=5).font  = F_MUTED
        r += 1

    def blank():
        nonlocal r
        r += 1

    # ── OCCUPANCY OVERVIEW (rows 3–10) ────────────────────────────────
    # Row 5 col 3 = total rent — must tie to RR Input row 124 col 2
    sec('OCCUPANCY & RENT OVERVIEW')
    hdr('Metric', '', 'Value', '', 'Notes', '')

    # Row 5: Total Actual Rent (test: row 5 col 3 = RR Input row 124 col 2)
    row('Total Actual Rent',   f"='{_RR}'!B124", NF_DOLLAR, 'Ties to RR Input total')
    row('Total Market Rent',   f"='{_RR}'!C124", NF_DOLLAR)
    row('Total Units',          prop['units'],    '#,##0')
    row('Occupied Units',      f"=COUNTIF('{_RR}'!D2:D{prop['units']+1},\">0\")", '#,##0')
    row('Vacancy Count',       f"={prop['units']}-D{r-1}", '#,##0',
        note='Uses cell above')
    row('Occupancy Rate',      f"=IF({prop['units']}>0,D{r-2}/{prop['units']},0)", NF_PCT)
    blank()

    # ── BY UNIT TYPE ──────────────────────────────────────────────────
    sec('BY UNIT TYPE')
    hdr('Type', 'Units', 'Avg Actual Rent', 'Avg Market Rent', 'LTM Gap', 'Occ %')

    # RR Input has columns: Unit, ActualRent, MarketRent, ...
    # We'll summarize by floor plan prefix (1BR/2BR)
    for unit_type, v in um.items():
        fill = P_ALT if r % 2 == 0 else P_WHITE
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=1).value = unit_type
        ws.cell(row=r, column=1).font  = F_LABEL
        ws.cell(row=r, column=2).value = v['count']
        ws.cell(row=r, column=2).font  = F_BODY
        ws.cell(row=r, column=3).value = v['market_rent']   # static fallback
        ws.cell(row=r, column=3).number_format = NF_DOLLAR
        ws.cell(row=r, column=4).value = v['market_rent']
        ws.cell(row=r, column=4).number_format = NF_DOLLAR
        ws.cell(row=r, column=5).value = f'=D{r}-C{r}'
        ws.cell(row=r, column=5).number_format = NF_DOLLAR
        r += 1

    blank()

    # ── ESCROW NAMES REFERENCE ────────────────────────────────────────
    sec('ESCROW ACCOUNTS')
    hdr('Account', '', 'Name', '', '', '')
    for i, name in enumerate(cfg['escrow_names'], 1):
        fill = P_ALT if r % 2 == 0 else P_WHITE
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=1).value = i
        ws.cell(row=r, column=1).font  = F_MUTED
        ws.cell(row=r, column=3).value = name
        ws.cell(row=r, column=3).font  = F_BODY
        r += 1

    last_row = r - 1

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 12

    hide_beyond(ws, last_row, MAX_COL)
    return ws
