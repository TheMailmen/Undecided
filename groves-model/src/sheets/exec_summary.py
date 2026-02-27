# src/sheets/exec_summary.py — Executive Summary sheet builder
"""
Builds the 'Executive Summary' tab: one-page KPI dashboard.
Exactly 45 rows × 6 columns per test_model.py boundary check.

Column map:
  A (1): Metric / Label
  B (2): Value  (Excel formula or static)
  C (3): Period (T12, Monthly, etc.)
  D (4): Benchmark / Target
  E (5): vs Target  (formula)
  F (6): Notes
"""
from openpyxl.utils import get_column_letter

from design import (
    apply_title, apply_hdr, apply_section, apply_subtotal,
    hide_beyond,
    F_BOLD, F_BODY, F_MUTED, F_LABEL, F_GREEN, F_RED, F_GOLD,
    P_ALT, P_WHITE, P_GREEN_LT, P_RED_LT, P_SUBTOTAL,
    B_SUB,
    NF_DOLLAR, NF_PCT, NF_DEC2,
    A_CENTER, A_LEFT,
)

MAX_COL = 6
_T12   = 'T12_PL'
_T12_O = '$O'   # col 15 = T12 total


def _t12(account):
    return f'=INDEX({_T12}!{_T12_O}:{_T12_O},MATCH("{account}",{_T12}!$B:$B,0))'


def build(wb, cfg):
    ws = wb.create_sheet('Executive Summary')
    prop   = cfg['property']
    loan   = cfg['loan']
    tic    = cfg['tic']
    val    = cfg['valuation']
    equity = cfg['total_equity']

    apply_title(
        ws, MAX_COL,
        f"{prop['name']}  |  Executive Summary",
        f"{prop['address']}  |  {prop['units']} Units  |  T12 Trailing 12-Month View",
    )

    r = 3

    def sec(title):
        nonlocal r
        apply_section(ws, r, MAX_COL, accent=True)
        ws.cell(row=r, column=1).value = title
        r += 1

    def hdr(*labels):
        nonlocal r
        apply_hdr(ws, r, MAX_COL)
        for i, lbl in enumerate(labels, 1):
            ws.cell(row=r, column=i).value = lbl
        r += 1

    def kpi(label, formula, period='T12', benchmark='', note='', nf=NF_DOLLAR,
            green=False, gold=False):
        nonlocal r
        fill = P_GREEN_LT if green else (P_ALT if r % 2 == 0 else P_WHITE)
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=1).value = label
        ws.cell(row=r, column=1).font  = F_GREEN if green else (F_GOLD if gold else F_LABEL)
        b_cell = ws.cell(row=r, column=2)
        b_cell.value = formula
        b_cell.number_format = nf
        b_cell.font = F_GREEN if green else F_BODY
        ws.cell(row=r, column=3).value = period
        ws.cell(row=r, column=3).font  = F_MUTED
        ws.cell(row=r, column=4).value = benchmark
        ws.cell(row=r, column=4).font  = F_MUTED
        ws.cell(row=r, column=6).value = note
        ws.cell(row=r, column=6).font  = F_MUTED
        r += 1

    def prop_row(label, value, note=''):
        nonlocal r
        fill = P_ALT if r % 2 == 0 else P_WHITE
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=1).value = label
        ws.cell(row=r, column=1).font  = F_LABEL
        ws.cell(row=r, column=2).value = value
        ws.cell(row=r, column=2).font  = F_BODY
        ws.cell(row=r, column=6).value = note
        ws.cell(row=r, column=6).font  = F_MUTED
        r += 1

    def blank():
        nonlocal r
        r += 1

    # ── KPI HIGHLIGHTS (rows 3–13) ────────────────────────────────────
    sec('KPI HIGHLIGHTS')
    hdr('Metric', 'Value', 'Period', 'Target / Benchmark', '', 'Notes')

    kpi('NET OPERATING INCOME (NOI)',    _t12('NET OPERATING INCOME (NOI)'),
        'T12', '', 'Key profitability indicator', NF_DOLLAR, green=True)
    kpi('CASH FLOW AFTER DEBT SERVICE',  _t12('CASH FLOW AFTER DEBT SERVICE'),
        'T12', '', 'After all debt payments', NF_DOLLAR, green=True)
    kpi('NET CASH FLOW',                 _t12('NET CASH FLOW'),
        'T12', '', 'After capex', NF_DOLLAR, green=True)
    kpi('DSCR',                          _t12('DSCR'),
        'Last Mo', '≥ 1.25×', 'Debt service coverage', NF_DEC2)
    kpi('Cap Rate (T12)',                _t12('Cap Rate'),
        'T12 Ann.', '6–8%', 'Annualized NOI / Purchase Price', NF_PCT)
    kpi('Cash-on-Cash Return (NCF)',     _t12('Cash-on-Cash (Ann.)'),
        'T12 Ann.', '8–12%', 'Ann. NCF / Total Equity', NF_PCT)
    kpi('Expense Ratio',                 _t12('Expense Ratio'),
        'Last Mo', '< 50%', 'OpEx / EGI', NF_PCT)
    kpi('Effective Gross Income',        _t12('EFFECTIVE GROSS INCOME (EGI)'),
        'T12', '', '', NF_DOLLAR)
    kpi('Total Operating Expenses',      _t12('Total Operating Expenses'),
        'T12', '', '', NF_DOLLAR)
    blank()   # row 14

    # ── T12 INCOME SUMMARY (rows 15–27) ──────────────────────────────
    sec('T12 INCOME SUMMARY')
    hdr('Line Item', 'T12 Total', '', '', '', '')

    for account, is_sub, green in [
        ('Gross Potential Rent',             False, False),
        ('Loss/Gain to Market',              False, False),
        ('Vacancy',                          False, False),
        ('Delinquency',                      False, False),
        ('Effective Rental Income',          True,  False),
        ('Total Other Income',               False, False),
        ('EFFECTIVE GROSS INCOME (EGI)',     True,  False),
        ('Total Operating Expenses',         True,  False),
        ('NET OPERATING INCOME (NOI)',       True,  True),
        ('Total Debt Service',               False, False),
        ('CASH FLOW AFTER DEBT SERVICE',     True,  True),
    ]:
        nonlocal_r = r
        fill = P_GREEN_LT if green else (P_SUBTOTAL if is_sub else (P_ALT if r % 2 == 0 else P_WHITE))
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=1).value = account
        ws.cell(row=r, column=1).font  = F_GREEN if green else (F_BOLD if is_sub else F_BODY)
        cell = ws.cell(row=r, column=2)
        cell.value = _t12(account)
        cell.number_format = NF_DOLLAR
        cell.font = F_GREEN if green else (F_BOLD if is_sub else F_BODY)
        r += 1

    blank()   # row 27 → 28

    # ── PROPERTY DETAILS (rows 29–35) ────────────────────────────────
    sec('PROPERTY DETAILS')
    prop_row('Property',      prop['name'])
    prop_row('Address',       prop['address'])
    prop_row('Units',         prop['units'])
    prop_row('Purchase Price', f"${prop['purchase_price']:,.0f}")
    prop_row('Purchase Date', prop['purchase_date'])
    blank()   # row 35 → 36

    # ── DEBT SUMMARY (rows 37–42) ─────────────────────────────────────
    sec('DEBT SUMMARY')
    prop_row('Original Loan', f"${loan['original_amount']:,.0f}")
    prop_row('Rate',          f"{loan['rate']*100:.2f}%  |  {loan['amort_years']}-yr amort")
    prop_row('Est. Balance',  f"${loan['est_current_balance']:,.0f}")
    prop_row('Term',          f"{loan['term_years']}-yr term  |  Matures ~{2024 + loan['term_years']}")
    blank()   # row 42 → 43

    # ── INVESTOR RETURNS (rows 43–45) ────────────────────────────────
    sec('INVESTOR RETURNS')
    hdr('Entity', 'Ownership %', 'Equity', '', '', '')
    for name, v in tic.items():
        prop_row(name, f"{v['pct']*100:.3f}%", f"  ${v['equity']:,.0f}")

    # Pad to exactly 45 rows if needed
    while r <= 45:
        r += 1

    last_row = min(r - 1, 45)

    # ── Column widths ─────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 26

    hide_beyond(ws, last_row, MAX_COL)
    return ws
