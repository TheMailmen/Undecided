# src/sheets/refi_stress.py — Refi Stress Test sheet builder
"""
Builds the 'Refi Stress Test' tab: refinance scenario analysis at
varying property values, LTVs, and interest rates.
"""
from openpyxl.utils import get_column_letter

from design import (
    apply_title, apply_hdr, apply_section, apply_subtotal,
    hide_beyond, input_cell,
    F_BOLD, F_BODY, F_LABEL, F_MUTED, F_GREEN, F_RED,
    P_ALT, P_WHITE, P_GREEN_LT, P_RED_LT, P_SUBTOTAL,
    NF_DOLLAR, NF_PCT, NF_DEC2,
    A_CENTER,
)

MAX_COL = 9   # label + 8 scenario columns


def build(wb, cfg):
    ws = wb.create_sheet('Refi Stress Test')
    prop  = cfg['property']
    refi  = cfg['refi']
    loan  = cfg['loan']
    equity = cfg['total_equity']

    apply_title(
        ws, MAX_COL,
        f"{prop['name']}  |  Refinance Stress Test",
        'Scenario analysis: value × LTV × rate combinations',
    )

    r = 3

    def sec(title):
        nonlocal r
        apply_section(ws, r, MAX_COL, accent=True)
        ws.cell(row=r, column=1).value = title
        r += 1

    def hdr(*labels):
        nonlocal r
        apply_hdr(ws, r, MAX_COL)
        for i, lbl in enumerate(labels, 1):
            ws.cell(row=r, column=i).value = lbl
        r += 1

    def row(label, vals, nf=NF_DOLLAR, bold=False, green=False):
        nonlocal r
        fill = P_GREEN_LT if green else (P_ALT if r % 2 == 0 else P_WHITE)
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=1).value = label
        ws.cell(row=r, column=1).font  = F_GREEN if green else (F_BOLD if bold else F_LABEL)
        for i, v in enumerate(vals, 2):
            cell = ws.cell(row=r, column=i)
            cell.value = v
            cell.number_format = nf
            cell.font = F_GREEN if green else (F_BOLD if bold else F_BODY)
        r += 1

    def blank():
        nonlocal r
        r += 1

    # ── ASSUMPTIONS ───────────────────────────────────────────────────
    sec('BASE ASSUMPTIONS')
    hdr('Input', 'Current', 'Refi Target', '', '', '', '', '', '')

    row('Property Value',     [prop['purchase_price'], refi['property_value']], NF_DOLLAR)
    row('LTV',                [0.0, refi['ltv']],   NF_PCT)
    row('Interest Rate',      [loan['rate'], refi['rate']], NF_PCT)
    row('Amortization',       [loan['amort_years'], refi['amort_years']])
    row('I/O',                ['No', 'Yes' if refi['io'] else 'No'])
    row('Current Loan Bal.',  [loan['est_current_balance'], ''], NF_DOLLAR)
    blank()

    # ── RATE SENSITIVITY (rows: rates × cols: property values) ────────
    sec('SENSITIVITY: NET PROCEEDS BY VALUE & RATE')
    # Header: rates across columns
    values = [
        refi['property_value'] * 0.85,
        refi['property_value'] * 0.90,
        refi['property_value'] * 0.95,
        refi['property_value'],
        refi['property_value'] * 1.05,
        refi['property_value'] * 1.10,
        refi['property_value'] * 1.15,
    ]
    ltv = refi['ltv']
    cur_bal = loan['est_current_balance']

    hdr('Rate \\ Value',
        *[f'${int(v/1e6*10)/10:.1f}M' for v in values[:8]])

    rates = [0.045, 0.050, 0.055, refi['rate'], 0.065, 0.070]
    for rate in rates:
        net_proceeds = [max(v * ltv - cur_bal, 0) for v in values[:8]]
        rate_label = f"{rate*100:.2f}%"
        fill = P_GREEN_LT if rate == refi['rate'] else (P_ALT if r % 2 == 0 else P_WHITE)
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=1).value = rate_label
        ws.cell(row=r, column=1).font  = F_BOLD if rate == refi['rate'] else F_LABEL
        for i, np_val in enumerate(net_proceeds, 2):
            cell = ws.cell(row=r, column=i)
            cell.value = np_val
            cell.number_format = NF_DOLLAR
            cell.font = F_BOLD if rate == refi['rate'] else F_BODY
        r += 1
    blank()

    # ── REFI SUMMARY ──────────────────────────────────────────────────
    sec('REFI SCENARIO SUMMARY')
    hdr('Metric', 'Value', 'Notes', '', '', '', '', '', '')

    new_loan   = refi['property_value'] * refi['ltv']
    net_proc   = new_loan - cur_bal
    ann_int    = new_loan * refi['rate']
    monthly_pi = ann_int / 12 if refi['io'] else (
        new_loan * (refi['rate'] / 12) /
        (1 - (1 + refi['rate'] / 12) ** (-refi['amort_years'] * 12))
    )

    row('New Loan Amount',    [new_loan],  NF_DOLLAR, bold=True)
    row('Net Cash-Out Proceeds', [net_proc], NF_DOLLAR, green=True)
    row('Annual Interest',    [ann_int],   NF_DOLLAR)
    row('Monthly P&I',        [monthly_pi], NF_DOLLAR)
    row('New LTV',            [refi['ltv']], NF_PCT)

    last_row = r - 1

    ws.column_dimensions['A'].width = 20
    for i in range(2, MAX_COL + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    hide_beyond(ws, last_row, MAX_COL)
    return ws
