# src/sheets/trailing.py — Trailing Analysis sheet builder
"""
Builds the 'Trailing Analysis' tab: period-over-period P&L comparison.

Column layout:
  A  (1) : GL Code
  B  (2) : Account
  C  (3) : Current Month (most recent)
  D  (4) : Prior Month
  E  (5) : Δ Month-over-Month
  F  (6) : T3 Current (last 3 months)
  G  (7) : T3 Prior   (3 months before that)
  H  (8) : Δ T3
  I  (9) : T12 Current (last 12 months)
  J  (10): T12 Prior   (prior 12 months)
  K  (11): Δ T12
"""
from datetime import datetime
from openpyxl.utils import get_column_letter

from design import (
    apply_title, apply_hdr, apply_pl_formatting, hide_beyond,
    NF_DOLLAR, NF_PCT, NF_DEC2, NF_DATE,
    A_CENTER, F_BOLD, F_MUTED,
)

MAX_COL = 11
_METRIC_PCT   = {'Cap Rate', 'Expense Ratio', 'Cash-on-Cash (CFADS)', 'Cash-on-Cash (Ann.)'}
_METRIC_RATIO = {'DSCR'}


def _sumif_for_months(month_dates, row_r):
    """Build a SUMIFS formula summing qPL_Fact for a list of date values."""
    if not month_dates:
        return '=0'
    parts = '+'.join(
        f'SUMIFS(qPL_Fact!$D:$D,qPL_Fact!$C:$C,$B{row_r},'
        f'qPL_Fact!$A:$A,DATE({m.year},{m.month},{m.day}))'
        for m in month_dates
    )
    return f'={parts}'


def build(wb, cfg):
    ws = wb.create_sheet('Trailing Analysis')

    # ── Collect sorted months from qPL_Fact ──────────────────────────
    qpl_ws = wb['qPL_Fact']
    seen, all_months = set(), []
    for row in qpl_ws.iter_rows(min_row=2, values_only=True):
        m = row[0]
        if m and m not in seen:
            seen.add(m)
            if hasattr(m, 'year'):
                all_months.append(m)
            else:
                all_months.append(datetime.strptime(str(m), '%Y-%m-%d'))
    all_months.sort()

    n = len(all_months)

    # Period windows (fallback gracefully if fewer months available)
    cur_mo   = all_months[-1:]                      # current month
    pri_mo   = all_months[-2:-1] if n >= 2 else []  # prior month
    t3_cur   = all_months[-3:]   if n >= 3 else all_months
    t3_pri   = all_months[-6:-3] if n >= 6 else (all_months[:3] if n >= 3 else all_months)
    t12_cur  = all_months[-12:]  if n >= 12 else all_months
    t12_pri  = all_months[-24:-12] if n >= 24 else (all_months[:12] if n >= 12 else all_months)

    # Format month labels
    def mo_label(months):
        if not months:
            return ''
        if len(months) == 1:
            return months[0].strftime('%b-%y')
        return f"{months[0].strftime('%b-%y')}–{months[-1].strftime('%b-%y')}"

    prop = cfg['property']
    apply_title(
        ws, MAX_COL,
        f"{prop['name']}  |  Trailing Analysis",
        f"Period-over-period variance  |  Current month: {mo_label(cur_mo)}",
    )

    # ── Column headers (row 3) ────────────────────────────────────────
    apply_hdr(ws, 3, MAX_COL)
    headers = [
        (1, 'GL'),
        (2, 'Account'),
        (3, mo_label(cur_mo)),
        (4, mo_label(pri_mo)),
        (5, 'Δ MoM'),
        (6, mo_label(t3_cur)),
        (7, mo_label(t3_pri)),
        (8, 'Δ T3'),
        (9, mo_label(t12_cur)),
        (10, mo_label(t12_pri)),
        (11, 'Δ T12'),
    ]
    for col, lbl in headers:
        cell = ws.cell(row=3, column=col)
        cell.value = lbl
        cell.alignment = A_CENTER

    # ── COA rows ─────────────────────────────────────────────────────
    coa        = cfg['coa']
    data_start = 4
    r          = data_start

    for gl, acct, rtype in coa:
        ws.cell(row=r, column=1).value = gl
        ws.cell(row=r, column=2).value = acct

        if rtype == 'spacer' or acct is None:
            r += 1
            continue

        if acct in _METRIC_PCT:
            nf = NF_PCT
        elif acct in _METRIC_RATIO:
            nf = NF_DEC2
        else:
            nf = NF_DOLLAR

        # Period values (cols 3–4, 6–7, 9–10)
        for col, months in [
            (3, cur_mo), (4, pri_mo),
            (6, t3_cur), (7, t3_pri),
            (9, t12_cur), (10, t12_pri),
        ]:
            cell = ws.cell(row=r, column=col)
            cell.value = _sumif_for_months(months, r)
            cell.number_format = nf

        # Delta cols (5 = MoM, 8 = T3, 11 = T12)
        for delta_col, cur_col, pri_col in [(5, 3, 4), (8, 6, 7), (11, 9, 10)]:
            cc = get_column_letter(cur_col)
            pc = get_column_letter(pri_col)
            d_cell = ws.cell(row=r, column=delta_col)
            d_cell.value = f'={cc}{r}-{pc}{r}'
            d_cell.number_format = nf

        r += 1

    last_row = r - 1
    apply_pl_formatting(ws, MAX_COL, data_start=data_start, data_end=last_row)

    # ── Column widths ─────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 30
    for col in range(3, MAX_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13

    ws.freeze_panes = 'C4'
    hide_beyond(ws, last_row, MAX_COL)
    return ws
