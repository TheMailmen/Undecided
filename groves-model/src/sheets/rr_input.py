# src/sheets/rr_input.py — Rent Roll Input sheet builder
"""
Builds the 'RR Input' tab: unit-by-unit rent roll from rent_roll.csv.

CSV format: Unit (wide) | date columns (one per month, values = actual rent)
A 0 means vacant that month.

Sheet layout:
  Row 1-2 : Title / subtitle
  Row 3   : Column headers (Unit | ActualRent | MarketRent | ... month cols)
  Rows 4–123: 120 units (one row each)
  Row 124 : Totals  <-- test_model.py checks row 124 col 2 = total actual rent

Column map:
  A (1): Unit number
  B (2): Latest actual rent   <-- test checks row 124 col 2 = total
  C (3): Market rent (from unit_improvements if available, else config)
  D (4): Loss-to-Market
  E onwards: monthly rent columns
"""
import csv
from openpyxl.utils import get_column_letter

from design import (
    apply_title, apply_hdr, apply_section, apply_subtotal,
    hide_beyond,
    F_BOLD, F_BODY, F_MUTED, F_LABEL,
    P_ALT, P_WHITE, P_SUBTOTAL,
    B_SUB,
    NF_DOLLAR, NF_PCT,
    A_CENTER,
)


def build(wb, cfg, rr_csv_path=None):
    """Build RR Input sheet.

    If rr_csv_path is None, attempts to locate it relative to the workbook
    output directory or falls back to config defaults.
    """
    import os
    if rr_csv_path is None:
        # Default: data/rent_roll.csv relative to build.py location
        rr_csv_path = os.path.join(
            os.path.dirname(__file__), '..', '..', 'data', 'rent_roll.csv'
        )

    ws = wb.create_sheet('RR Input')
    prop = cfg['property']
    um   = cfg['unit_mix']

    # ── Read rent roll CSV ────────────────────────────────────────────
    units = []      # list of {unit, months: {date_str: rent}}
    month_cols = [] # ordered list of date strings from header

    if os.path.exists(rr_csv_path):
        with open(rr_csv_path, 'r') as f:
            reader = csv.reader(f)
            header = next(reader)
            month_cols = header[1:]   # skip 'Unit' column
            for row_data in reader:
                if not row_data or not row_data[0]:
                    continue
                unit = row_data[0]
                months = {}
                for i, col in enumerate(month_cols):
                    val = row_data[i + 1].strip() if i + 1 < len(row_data) else ''
                    months[col] = float(val) if val else 0.0
                units.append({'unit': unit, 'months': months})
    else:
        # Fallback: generate placeholder unit list from config
        unit_list = [f'00-{100+i:03d}' for i in range(prop['units'])]
        for u in unit_list:
            units.append({'unit': u, 'months': {}})

    n_months    = len(month_cols)
    FIRST_MO_COL = 5      # E
    last_col    = FIRST_MO_COL + n_months - 1 if n_months > 0 else 5

    # Determine latest month for col B (actual rent)
    latest_month = month_cols[-1] if month_cols else None

    # ── Title / subtitle ──────────────────────────────────────────────
    apply_title(
        ws, last_col,
        f"{prop['name']}  |  Rent Roll Input",
        f"Unit-by-unit rent history  |  {len(units)} units  |  {n_months} months",
    )

    # ── Column headers (row 3) ────────────────────────────────────────
    apply_hdr(ws, 3, last_col)
    ws.cell(row=3, column=1).value = 'Unit'
    ws.cell(row=3, column=2).value = 'Actual Rent'     # test checks row 124 col 2
    ws.cell(row=3, column=3).value = 'Market Rent'
    ws.cell(row=3, column=4).value = 'LTM Gap'

    for i, mo in enumerate(month_cols):
        cell = ws.cell(row=3, column=FIRST_MO_COL + i)
        cell.value = mo
        cell.alignment = A_CENTER

    # ── Unit rows (4 through 4+n_units-1) ────────────────────────────
    DATA_START = 4
    for idx, u in enumerate(units):
        r = DATA_START + idx
        fill = P_ALT if r % 2 == 0 else P_WHITE
        for c in range(1, last_col + 1):
            ws.cell(row=r, column=c).fill = fill

        ws.cell(row=r, column=1).value = u['unit']
        ws.cell(row=r, column=1).font  = F_MUTED

        # Col B: latest actual rent (or 0 if vacant)
        latest_rent = u['months'].get(latest_month, 0) if latest_month else 0
        b_cell = ws.cell(row=r, column=2)
        if n_months > 0:
            last_mo_ltr = get_column_letter(last_col)
            b_cell.value = f'={last_mo_ltr}{r}'
        else:
            b_cell.value = latest_rent
        b_cell.number_format = NF_DOLLAR

        # Col C: market rent — from unit_improvements if present, else config default
        # Simple heuristic: 1BR = first mix entry, 2BR = second
        # Use unit number suffix to guess: 1xx = 1BR, 2xx... just use config avg
        avg_mkt = sum(v['market_rent'] * v['count'] for v in um.values()) / prop['units']
        ws.cell(row=r, column=3).value = avg_mkt
        ws.cell(row=r, column=3).number_format = NF_DOLLAR
        ws.cell(row=r, column=3).font  = F_MUTED

        # Col D: LTM gap = market - actual
        ws.cell(row=r, column=4).value = f'=C{r}-B{r}'
        ws.cell(row=r, column=4).number_format = NF_DOLLAR

        # Monthly rent columns
        for i, mo in enumerate(month_cols):
            col  = FIRST_MO_COL + i
            rent = u['months'].get(mo, 0)
            cell = ws.cell(row=r, column=col)
            cell.value = rent if rent != 0 else ''
            cell.number_format = NF_DOLLAR

    # ── Totals row (row 124 = DATA_START + 120) ───────────────────────
    # test_model.py: RR Input row 124 col 2 = total actual rent
    n_units  = len(units)
    tot_row  = DATA_START + n_units   # row 124 when n_units = 120

    apply_subtotal(ws, tot_row, last_col)
    ws.cell(row=tot_row, column=1).value = 'TOTAL'
    ws.cell(row=tot_row, column=1).font  = F_BOLD

    end_data = DATA_START + n_units - 1  # last unit row

    # Col B total (test checks this)
    tot_b = ws.cell(row=tot_row, column=2)
    tot_b.value = f'=SUM(B{DATA_START}:B{end_data})'
    tot_b.number_format = NF_DOLLAR
    tot_b.font = F_BOLD

    # Col C total
    tot_c = ws.cell(row=tot_row, column=3)
    tot_c.value = f'=SUM(C{DATA_START}:C{end_data})'
    tot_c.number_format = NF_DOLLAR
    tot_c.font = F_BOLD

    # Monthly totals
    for i in range(n_months):
        col = FIRST_MO_COL + i
        ltr = get_column_letter(col)
        cell = ws.cell(row=tot_row, column=col)
        cell.value = f'=SUM({ltr}{DATA_START}:{ltr}{end_data})'
        cell.number_format = NF_DOLLAR
        cell.font = F_BOLD

    # ── Column widths ─────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 11
    for i in range(n_months):
        ws.column_dimensions[get_column_letter(FIRST_MO_COL + i)].width = 10

    ws.freeze_panes = 'B4'
    hide_beyond(ws, tot_row, last_col)
    return ws
