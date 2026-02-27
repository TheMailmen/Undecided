# src/sheets/distribution.py — Distribution Model sheet builder
"""
Builds the 'Distribution_Model' tab: TIC investor distribution waterfall.
Maps T12 cash flows to each entity by ownership percentage.
Font limit: 15 (vs 10 for other sheets) per test_model.py.
"""
from openpyxl.utils import get_column_letter

from design import (
    apply_title, apply_hdr, apply_section, apply_subtotal,
    hide_beyond, input_cell,
    F_BOLD, F_BODY, F_LABEL, F_MUTED, F_GREEN, F_GOLD,
    P_ALT, P_WHITE, P_GREEN_LT, P_SUBTOTAL,
    B_SUB,
    NF_DOLLAR, NF_PCT, NF_DEC2,
    A_CENTER, A_LEFT,
)

MAX_COL = 8
_T12   = 'T12_PL'
_T12_O = '$O'


def _t12(account):
    return f'=INDEX({_T12}!{_T12_O}:{_T12_O},MATCH("{account}",{_T12}!$B:$B,0))'


def build(wb, cfg):
    ws = wb.create_sheet('Distribution_Model')

    prop   = cfg['property']
    tic    = cfg['tic']
    equity = cfg['total_equity']
    refi   = cfg['refi']

    apply_title(
        ws, MAX_COL,
        f"{prop['name']}  |  Distribution Model",
        'TIC Investor Waterfall — T12 Trailing Actuals',
    )

    r = 3

    def sec(title):
        nonlocal r
        apply_section(ws, r, MAX_COL, accent=True)
        ws.cell(row=r, column=1).value = title
        r += 1

    def hdr(*labels):
        nonlocal r
        apply_hdr(ws, r, MAX_COL)
        for i, lbl in enumerate(labels, 1):
            ws.cell(row=r, column=i).value = lbl
        r += 1

    def data_row(label, col2_val='', col2_nf=NF_DOLLAR, note='',
                 bold=False, green=False, is_sub=False):
        nonlocal r
        fill = P_GREEN_LT if green else (P_SUBTOTAL if is_sub else (P_ALT if r % 2 == 0 else P_WHITE))
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=1).value = label
        ws.cell(row=r, column=1).font  = F_GREEN if green else (F_BOLD if (bold or is_sub) else F_LABEL)
        if col2_val != '':
            cell = ws.cell(row=r, column=2)
            cell.value = col2_val
            cell.number_format = col2_nf
            cell.font = F_GREEN if green else (F_BOLD if (bold or is_sub) else F_BODY)
        if note:
            ws.cell(row=r, column=3).value = note
            ws.cell(row=r, column=3).font  = F_MUTED
        r += 1

    def blank():
        nonlocal r
        r += 1

    # ── CASH FLOW SUMMARY (source data) ──────────────────────────────
    sec('CASH FLOW SUMMARY — T12 ACTUALS')
    hdr('Line Item', 'T12 Total', 'Notes', '', '', '', '', '')

    cf_rows = [
        ('EFFECTIVE GROSS INCOME (EGI)',  False, False),
        ('Total Operating Expenses',       False, False),
        ('NET OPERATING INCOME (NOI)',     True,  True),
        ('Total Debt Service',             False, False),
        ('CASH FLOW AFTER DEBT SERVICE',   True,  True),
        ('Total Capital Expenditures',     False, False),
        ('NET CASH FLOW',                  True,  True),
    ]
    for account, is_sub, green in cf_rows:
        data_row(account, _t12(account), NF_DOLLAR, is_sub=is_sub, green=green)

    blank()

    # ── DISTRIBUTION WATERFALL ────────────────────────────────────────
    sec('DISTRIBUTION WATERFALL')
    hdr('', 'Amount', '% of NCF', 'Notes', '', '', '', '')

    # Track the NCF row for formula reference
    ncf_row = r
    data_row('Available for Distribution (NCF)', _t12('NET CASH FLOW'), NF_DOLLAR,
             bold=True, green=True)
    blank()

    # Preferred return tier
    data_row('Preferred Return (NOAH)',
             f'=MIN(C{r-2}*{refi["noah_equity_pct"]},B{ncf_row}*{refi["noah_pref_return"]})',
             NF_DOLLAR, note=f"{refi['noah_pref_return']*100:.2f}% pref on {refi['noah_equity_pct']*100:.0f}% equity")
    pref_row = r - 1
    data_row('Remaining After Pref',
             f'=MAX(B{ncf_row}-B{pref_row},0)', NF_DOLLAR, is_sub=True)
    remaining_row = r - 1
    blank()

    # ── PER-ENTITY ALLOCATIONS ────────────────────────────────────────
    sec('PER-ENTITY ALLOCATIONS')
    entity_names = list(tic.keys())
    hdr('Entity', 'Ownership %', 'Equity', 'NCF Share', 'Ann. CoC', 'Est. Value', '', '')

    entity_rows = {}
    for name, v in tic.items():
        fill = P_ALT if r % 2 == 0 else P_WHITE
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill

        ws.cell(row=r, column=1).value = name
        ws.cell(row=r, column=1).font  = F_BOLD

        # Col B: ownership %
        input_cell(ws, r, 2, value=v['pct'], nf=NF_PCT)
        # Col C: equity
        input_cell(ws, r, 3, value=v['equity'], nf=NF_DOLLAR)
        # Col D: NCF share = NCF * ownership %
        d = ws.cell(row=r, column=4)
        d.value = f'=B{ncf_row}*B{r}'
        d.number_format = NF_DOLLAR
        # Col E: Annual CoC = NCF share / equity
        e = ws.cell(row=r, column=5)
        e.value = f'=IF(C{r}>0,D{r}/C{r},0)'
        e.number_format = NF_PCT
        # Col F: Est. value = BOV * ownership %
        f_cell = ws.cell(row=r, column=6)
        f_cell.value = f'={cfg["valuation"]["bov_mid"]}*B{r}'
        f_cell.number_format = NF_DOLLAR

        entity_rows[name] = r
        r += 1

    # Totals row
    apply_subtotal(ws, r, MAX_COL)
    ws.cell(row=r, column=1).value = 'TOTAL'
    ws.cell(row=r, column=1).font  = F_BOLD
    for col, nf in [(2, NF_PCT), (3, NF_DOLLAR), (4, NF_DOLLAR), (6, NF_DOLLAR)]:
        ltr = get_column_letter(col)
        start = min(entity_rows.values())
        end   = max(entity_rows.values())
        cell  = ws.cell(row=r, column=col)
        cell.value = f'=SUM({ltr}{start}:{ltr}{end})'
        cell.number_format = nf
        cell.font = F_BOLD
    ws.cell(row=r, column=5).value = f'=IF(C{r}>0,D{r}/C{r},0)'
    ws.cell(row=r, column=5).number_format = NF_PCT
    ws.cell(row=r, column=5).font = F_BOLD
    r += 1
    blank()

    # ── ASSUMPTIONS ──────────────────────────────────────────────────
    sec('KEY ASSUMPTIONS')
    data_row('Purchase Price',   prop['purchase_price'],   NF_DOLLAR)
    data_row('Total Equity',     equity,                   NF_DOLLAR)
    data_row('BOV Midpoint',     cfg['valuation']['bov_mid'], NF_DOLLAR)
    data_row('NOAH Pref Return', refi['noah_pref_return'], NF_PCT)

    last_row = r - 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    hide_beyond(ws, last_row, MAX_COL)
    return ws
