# src/sheets/escrow_summary.py — Escrow Summary sheet builder
"""
Builds the 'Escrow_Summary' tab: cumulative running balance for each
escrow/reserve account, derived from escrow_input data.
"""
from openpyxl.utils import get_column_letter

from design import (
    apply_title, apply_hdr, apply_section, apply_subtotal,
    hide_beyond,
    F_BOLD, F_BODY, F_LABEL, F_MUTED, F_GREEN,
    P_ALT, P_WHITE, P_SUBTOTAL, P_GREEN_LT,
    NF_DOLLAR, NF_DATE,
    A_CENTER,
)

MAX_COL = 6
_ESC = 'Escrow_Input'


def build(wb, cfg):
    ws = wb.create_sheet('Escrow_Summary')
    prop  = cfg['property']
    names = cfg['escrow_names']

    apply_title(
        ws, MAX_COL,
        f"{prop['name']}  |  Escrow Summary",
        'Cumulative balances for tax, insurance, and capital reserve accounts',
    )

    r = 3

    def sec(title):
        nonlocal r
        apply_section(ws, r, MAX_COL, accent=True)
        ws.cell(row=r, column=1).value = title
        r += 1

    def hdr(*labels):
        nonlocal r
        apply_hdr(ws, r, MAX_COL)
        for i, lbl in enumerate(labels, 1):
            ws.cell(row=r, column=i).value = lbl
        r += 1

    def data_row(label, vals=None, nf=NF_DOLLAR, is_sub=False, green=False):
        nonlocal r
        from design import P_SUBTOTAL as SUB, P_GREEN_LT as GRN
        fill = GRN if green else (SUB if is_sub else (P_ALT if r % 2 == 0 else P_WHITE))
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=1).value = label
        ws.cell(row=r, column=1).font  = F_GREEN if green else (F_BOLD if is_sub else F_LABEL)
        if vals:
            for col_idx, val in enumerate(vals, 2):
                cell = ws.cell(row=r, column=col_idx)
                cell.value = val
                cell.number_format = nf
                cell.font = F_GREEN if green else (F_BOLD if is_sub else F_BODY)
        r += 1

    def blank():
        nonlocal r
        r += 1

    # ── ACCOUNT BALANCES ──────────────────────────────────────────────
    sec('ESCROW ACCOUNT BALANCES')
    hdr('Account', 'Total Deposits', 'Total Payments', 'Net Balance', 'Notes', '')

    # Cross-reference Escrow_Input for SUMIF by account name
    for acct_name in names:
        fill = P_ALT if r % 2 == 0 else P_WHITE
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=1).value = acct_name
        ws.cell(row=r, column=1).font  = F_LABEL

        # Deposits = SUMIF on Escrow_Input col C where col B = account name
        dep = ws.cell(row=r, column=2)
        dep.value = f'=SUMIF(\'{_ESC}\'!$B:$B,A{r},\'{_ESC}\'!$C:$C)'
        dep.number_format = NF_DOLLAR

        # Payments
        pay = ws.cell(row=r, column=3)
        pay.value = f'=SUMIF(\'{_ESC}\'!$B:$B,A{r},\'{_ESC}\'!$D:$D)'
        pay.number_format = NF_DOLLAR

        # Net balance
        net = ws.cell(row=r, column=4)
        net.value = f'=B{r}-C{r}'
        net.number_format = NF_DOLLAR
        net.font = F_BOLD

        r += 1

    # Total row
    apply_subtotal(ws, r, MAX_COL, green=True)
    ws.cell(row=r, column=1).value = 'TOTAL ALL ACCOUNTS'
    ws.cell(row=r, column=1).font  = F_BOLD
    n = len(names)
    start = r - n
    for col, ltr in [(2, 'B'), (3, 'C'), (4, 'D')]:
        cell = ws.cell(row=r, column=col)
        cell.value = f'=SUM({ltr}{start}:{ltr}{r-1})'
        cell.number_format = NF_DOLLAR
        cell.font = F_BOLD
    r += 1
    blank()

    # ── MONTHLY DEPOSIT SUMMARY ───────────────────────────────────────
    sec('MONTHLY DEPOSITS BY ACCOUNT')
    hdr('Month', *names, 'Total', '')

    # Pull from Escrow_Input: SUMIFS by month and account
    # Escrow_Input col A = Month, col B = EscrowName, col C = Deposits
    # Get unique months from qPL_Fact
    qpl_ws = wb["qPL_Fact"]
    seen, months = set(), []
    for row in qpl_ws.iter_rows(min_row=2, values_only=True):
        m = row[0]
        if m and m not in seen:
            seen.add(m)
            months.append(m)
    months.sort()

    for m in months:
        fill = P_ALT if r % 2 == 0 else P_WHITE
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = fill

        # Month label
        ws.cell(row=r, column=1).value = m
        ws.cell(row=r, column=1).number_format = NF_DATE
        ws.cell(row=r, column=1).font = F_MUTED

        col_start = 2
        for i, acct_name in enumerate(names):
            col = col_start + i
            if col > MAX_COL - 1:
                break
            cell = ws.cell(row=r, column=col)
            if hasattr(m, 'year'):
                date_ref = f'DATE({m.year},{m.month},{m.day})'
            else:
                date_ref = f'"{m}"'
            cell.value = (
                f'=SUMIFS(\'{_ESC}\'!$C:$C,'
                f'\'{_ESC}\'!$A:$A,{date_ref},'
                f'\'{_ESC}\'!$B:$B,"{acct_name}")'
            )
            cell.number_format = NF_DOLLAR

        # Total for this month (sum of account cols)
        total_col = col_start + len(names)
        if total_col <= MAX_COL:
            start_ltr = get_column_letter(col_start)
            end_ltr   = get_column_letter(col_start + len(names) - 1)
            tot = ws.cell(row=r, column=total_col)
            tot.value = f'=SUM({start_ltr}{r}:{end_ltr}{r})'
            tot.number_format = NF_DOLLAR
            tot.font = F_BOLD

        r += 1

    last_row = r - 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 12

    hide_beyond(ws, last_row, MAX_COL)
    return ws
