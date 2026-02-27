# tests/test_model.py — Validates the built model
"""
Run after build.py to verify the model is correct.

Usage: python tests/test_model.py output/Groves_Investor_Model.xlsx
"""
import sys
import openpyxl
from openpyxl.utils import get_column_letter


def test_model(path):
    wb = openpyxl.load_workbook(path)
    wb_data = openpyxl.load_workbook(path, data_only=True)
    errors = []
    
    # ── Sheet existence ──
    required_visible = [
        'How To Use', 'Executive Summary', 'Assumptions', 'Full P&L',
        'T12_PL', 'Trailing Analysis', 'Distribution_Model', 'TIC Ownership',
        'RR Summary', 'RR Input', 'Unit Improvements', 'Escrow_Summary',
        'Escrow_Input', 'CapEx Profile', 'Refi Stress Test', 'Rent Comps',
    ]
    required_hidden = ['qPL_Fact']
    
    for name in required_visible:
        if name not in wb.sheetnames:
            errors.append(f"MISSING SHEET: {name}")
        elif wb[name].sheet_state != 'visible':
            errors.append(f"SHOULD BE VISIBLE: {name}")
    
    for name in required_hidden:
        if name not in wb.sheetnames:
            errors.append(f"MISSING ENGINE: {name}")
    
    # ── Title bar consistency ──
    for name in required_visible:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        fill = ws.cell(row=1, column=1).fill
        if fill.fgColor and hasattr(fill.fgColor, 'rgb'):
            rgb = fill.fgColor.rgb
            if not rgb.endswith('0D1B2A'):
                errors.append(f"TITLE FILL WRONG on {name}: #{rgb}")
    
    # ── Font count per sheet ──
    for name in required_visible:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        fonts = set()
        for row in range(1, min(ws.max_row + 1, 150)):
            for col in range(1, min(ws.max_column + 1, 25)):
                c = ws.cell(row=row, column=col)
                if c.value is not None:
                    fn = c.font.name or 'def'
                    fs = c.font.size or 0
                    fb = 'B' if c.font.bold else ''
                    fc = c.font.color.rgb if c.font.color and hasattr(c.font.color, 'rgb') and isinstance(c.font.color.rgb, str) else 'auto'
                    fonts.add(f"{fn}/{fs}/{fb}/{fc}")
        limit = 15 if name == 'Distribution_Model' else 10
        if len(fonts) > limit:
            errors.append(f"TOO MANY FONTS on {name}: {len(fonts)} (limit {limit})")
    
    # ── TIC ownership sums to 100% ──
    if 'TIC Ownership' in wb.sheetnames:
        ws = wb_data['TIC Ownership']
        total_pct = ws.cell(row=7, column=3).value
        if total_pct is not None and abs(total_pct - 1.0) > 0.001:
            errors.append(f"TIC OWNERSHIP doesn't sum to 100%: {total_pct}")
    
    # ── Cross-sheet NOI consistency ──
    if all(n in wb_data.sheetnames for n in ['TIC Ownership', 'T12_PL']):
        tic_noi = wb_data['TIC Ownership'].cell(row=21, column=3).value
        # T12 NOI = row where col B = 'NET OPERATING INCOME (NOI)', col O (T-12 Total)
        t12_noi = None
        for row in range(1, 90):
            if wb_data['T12_PL'].cell(row=row, column=2).value == 'NET OPERATING INCOME (NOI)':
                t12_noi = wb_data['T12_PL'].cell(row=row, column=15).value
                break
        if tic_noi and t12_noi and abs(tic_noi - t12_noi) > 1:
            errors.append(f"NOI MISMATCH: TIC={tic_noi:.0f} vs T12={t12_noi:.0f}")
    
    # ── RR Summary ties to RR Input ──
    if all(n in wb_data.sheetnames for n in ['RR Summary', 'RR Input']):
        rr_total = wb_data['RR Input'].cell(row=124, column=2).value  # First month total
        summary_total = wb_data['RR Summary'].cell(row=5, column=3).value  # First month
        if rr_total and summary_total and abs(rr_total - summary_total) > 1:
            errors.append(f"RR MISMATCH: Input={rr_total} vs Summary={summary_total}")
    
    # ── Hidden rows/cols beyond data ──
    from collections import Counter
    unhidden_beyond = Counter()
    boundaries = {
        'How To Use': (155, 5), 'Executive Summary': (45, 6),
        'Full P&L': (112, 24), 'TIC Ownership': (27, 8),
    }
    for name, (lr, lc) in boundaries.items():
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        for col_idx in range(lc + 1, 30):
            cl = get_column_letter(col_idx)
            if not ws.column_dimensions[cl].hidden:
                unhidden_beyond[name] += 1
    for name, count in unhidden_beyond.items():
        if count > 0:
            errors.append(f"UNHIDDEN COLS beyond data on {name}: {count}")
    
    # ── Formula check (no static where formula expected) ──
    if 'TIC Ownership' in wb.sheetnames:
        ws = wb['TIC Ownership']
        for r in [4, 5, 6]:
            for c in [5, 6, 7, 8]:
                v = ws.cell(row=r, column=c).value
                if not (isinstance(v, str) and v.startswith('=')):
                    errors.append(f"TIC row {r} col {chr(64+c)} should be formula, got: {type(v).__name__}")
    
    # ── Report ──
    print(f"\n{'='*50}")
    print(f"MODEL VALIDATION: {path}")
    print(f"{'='*50}")
    if errors:
        print(f"\n❌ {len(errors)} ISSUES FOUND:\n")
        for e in errors:
            print(f"  • {e}")
    else:
        print("\n✅ ALL CHECKS PASSED")
    print(f"\nSheets: {len(wb.sheetnames)} total, {sum(1 for s in wb.sheetnames if wb[s].sheet_state=='visible')} visible")
    
    return len(errors) == 0


if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else 'output/Groves_Investor_Model.xlsx'
    success = test_model(path)
    sys.exit(0 if success else 1)
