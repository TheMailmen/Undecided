# src/sheets/distribution.py — Monthly cash flow distribution waterfall
import csv
from datetime import datetime
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from design import (apply_title, apply_hdr, apply_section, apply_subtotal,
                     set_col_widths, hide_beyond, input_cell,
                     F_BODY, F_BOLD, F_LABEL, F_GREEN, F_MUTED, F_GOLD,
                     P_ALT, P_WHITE, P_GREEN_LT,
                     NF_DOLLAR, NF_PCT, NF_PCT2, NF_NUM, NF_DATE)


def build(wb, cfg):
    ws = wb.create_sheet('Distribution_Model')
    n = cfg['qpl_rows']

    # Get months from qPL_Fact
    qpl = wb['qPL_Fact']
    months_set = set()
    for r in range(2, n + 2):
        v = qpl.cell(row=r, column=1).value
        if v:
            if isinstance(v, datetime):
                months_set.add(v.strftime('%Y-%m-%d'))
            else:
                months_set.add(str(v))
    months = sorted(months_set)

    # Layout: Col A=spacer, Col B=Label, Cols 3..3+len(months)-1=monthly data,
    #         Then: spacer, T-12 Total
    first_data_col = 3
    last_month_col = first_data_col + len(months) - 1
    spacer_col = last_month_col + 1
    t12_col = spacer_col + 1
    max_col = t12_col

    apply_title(ws, max_col,
                'DISTRIBUTION MODEL',
                'Monthly cash flow waterfall from NOI to owner distributions.')

    # Headers
    ws.cell(row=3, column=1).value = ''
    ws.cell(row=3, column=2).value = 'Item'
    for i, m in enumerate(months):
        dt = datetime.strptime(m, '%Y-%m-%d')
        ws.cell(row=3, column=first_data_col + i).value = dt
        ws.cell(row=3, column=first_data_col + i).number_format = NF_DATE
    ws.cell(row=3, column=t12_col).value = 'T-12 Total'
    apply_hdr(ws, 3, max_col)

    from openpyxl.utils import get_column_letter
    widths = {'A': 4, 'B': 35}
    for c in range(first_data_col, last_month_col + 1):
        widths[get_column_letter(c)] = 12
    widths[get_column_letter(spacer_col)] = 2
    widths[get_column_letter(t12_col)] = 14
    set_col_widths(ws, widths)

    sr = f'qPL_Fact!$D$2:$D${n}'
    ac = f'qPL_Fact!$C$2:$C${n}'
    dtr = f'qPL_Fact!$A$2:$A${n}'

    def sumifs_month(account, month_str):
        dt = datetime.strptime(month_str, '%Y-%m-%d')
        return f'SUMIFS({sr},{ac},"{account}",{dtr},DATE({dt.year},{dt.month},1))'

    def write_row(row, label, formula_fn=None, is_green=False, is_subtotal=False,
                  is_section=False, nf=NF_DOLLAR, is_input=False, static_value=None):
        if is_section:
            apply_section(ws, row, max_col, accent=True)
            ws.cell(row=row, column=2).value = label
            return

        if is_subtotal or is_green:
            apply_subtotal(ws, row, max_col, green=is_green)
        else:
            fill = P_ALT if row % 2 == 0 else P_WHITE
            for c in range(1, max_col + 1):
                ws.cell(row=row, column=c).fill = fill

        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font = F_GREEN if is_green else (F_BOLD if is_subtotal else F_LABEL)

        if formula_fn:
            for i, m in enumerate(months):
                col = first_data_col + i
                ws.cell(row=row, column=col).value = f'={formula_fn(m)}'
                ws.cell(row=row, column=col).number_format = nf
            # T-12 total
            first = f'{get_column_letter(first_data_col)}{row}'
            last = f'{get_column_letter(last_month_col)}{row}'
            ws.cell(row=row, column=t12_col).value = f'=SUM({first}:{last})'
            ws.cell(row=row, column=t12_col).number_format = nf
        elif is_input and static_value is not None:
            for i in range(len(months)):
                col = first_data_col + i
                input_cell(ws, row, col, static_value, nf)
            first = f'{get_column_letter(first_data_col)}{row}'
            last = f'{get_column_letter(last_month_col)}{row}'
            ws.cell(row=row, column=t12_col).value = f'=SUM({first}:{last})'
            ws.cell(row=row, column=t12_col).number_format = nf

    row = 4
    # ── CASH FLOW SECTION ──
    write_row(row, 'CASH FLOW SOURCES', is_section=True); row += 1

    # NOI
    noi_row = row
    write_row(row, 'NET OPERATING INCOME (NOI)',
              lambda m: sumifs_month('NET OPERATING INCOME (NOI)', m),
              is_green=True); row += 1

    # Debt Service
    ds_row = row
    write_row(row, 'Total Debt Service',
              lambda m: sumifs_month('Total Debt Service', m)); row += 1

    # CFADS
    cfads_row = row
    write_row(row, 'CASH FLOW AFTER DEBT SERVICE',
              lambda m: sumifs_month('CASH FLOW AFTER DEBT SERVICE', m),
              is_green=True); row += 1

    # CapEx
    capex_row = row
    write_row(row, 'Total Capital Expenditures',
              lambda m: sumifs_month('Total Capital Expenditures', m)); row += 1

    # Net Cash Flow
    ncf_row = row
    write_row(row, 'NET CASH FLOW',
              lambda m: sumifs_month('NET CASH FLOW', m),
              is_green=True); row += 1

    row += 1  # spacer

    # ── DISTRIBUTION WATERFALL ──
    write_row(row, 'DISTRIBUTION WATERFALL', is_section=True); row += 1

    # Operating reserve / cushion (input)
    cushion_row = row
    write_row(row, 'Less: Operating Reserve', is_input=True, static_value=0); row += 1

    # Asset management fee
    am_row = row
    write_row(row, 'Less: Asset Mgmt Fee',
              lambda m: sumifs_month('Asset Mgmt Fee', m)); row += 1

    row += 1  # spacer

    # Free Cash for Distribution
    free_row = row
    apply_subtotal(ws, row, max_col, green=True)
    ws.cell(row=row, column=2).value = 'FREE CASH FOR DISTRIBUTION'
    ws.cell(row=row, column=2).font = F_GREEN
    for i in range(len(months)):
        col = first_data_col + i
        ws.cell(row=row, column=col).value = (
            f'={get_column_letter(col)}{ncf_row}'
            f'-{get_column_letter(col)}{cushion_row}'
            f'-{get_column_letter(col)}{am_row}')
        ws.cell(row=row, column=col).number_format = NF_DOLLAR
    first = f'{get_column_letter(first_data_col)}{row}'
    last = f'{get_column_letter(last_month_col)}{row}'
    ws.cell(row=row, column=t12_col).value = f'=SUM({first}:{last})'
    ws.cell(row=row, column=t12_col).number_format = NF_DOLLAR
    row += 1

    row += 1  # spacer

    # ── OWNER DISTRIBUTIONS ──
    write_row(row, 'OWNER DISTRIBUTIONS', is_section=True); row += 1

    # Headers for ownership
    own_hdr_row = row
    ws.cell(row=row, column=2).value = 'Owner'
    apply_hdr(ws, row, max_col)
    row += 1

    # Each TIC owner's distribution
    owner_rows = {}
    for owner, info in cfg['tic'].items():
        pct = info['pct']
        fill = P_ALT if row % 2 == 0 else P_WHITE
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=2).value = f'{owner} ({pct*100:.3f}%)'
        ws.cell(row=row, column=2).font = F_LABEL
        for i in range(len(months)):
            col = first_data_col + i
            ws.cell(row=row, column=col).value = f'={get_column_letter(col)}{free_row}*{pct}'
            ws.cell(row=row, column=col).number_format = NF_DOLLAR
        first = f'{get_column_letter(first_data_col)}{row}'
        last = f'{get_column_letter(last_month_col)}{row}'
        ws.cell(row=row, column=t12_col).value = f'=SUM({first}:{last})'
        ws.cell(row=row, column=t12_col).number_format = NF_DOLLAR
        owner_rows[owner] = row
        row += 1

    # Total distributions
    first_owner = min(owner_rows.values())
    last_owner = max(owner_rows.values())
    apply_subtotal(ws, row, max_col)
    ws.cell(row=row, column=2).value = 'Total Distributions'
    ws.cell(row=row, column=2).font = F_BOLD
    for i in range(len(months)):
        col = first_data_col + i
        ws.cell(row=row, column=col).value = (
            f'=SUM({get_column_letter(col)}{first_owner}:{get_column_letter(col)}{last_owner})')
        ws.cell(row=row, column=col).number_format = NF_DOLLAR
    first = f'{get_column_letter(first_data_col)}{row}'
    last = f'{get_column_letter(last_month_col)}{row}'
    ws.cell(row=row, column=t12_col).value = f'=SUM({first}:{last})'
    ws.cell(row=row, column=t12_col).number_format = NF_DOLLAR
    row += 1

    row += 1  # spacer

    # ── KEY METRICS ──
    write_row(row, 'KEY METRICS', is_section=True); row += 1

    # DSCR
    ws.cell(row=row, column=2).value = 'DSCR'
    ws.cell(row=row, column=2).font = F_LABEL
    fill = P_ALT if row % 2 == 0 else P_WHITE
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill
    for i, m in enumerate(months):
        col = first_data_col + i
        noi_f = sumifs_month('NET OPERATING INCOME (NOI)', m)
        ds_f = sumifs_month('Total Debt Service', m)
        ws.cell(row=row, column=col).value = f'=IF({ds_f}<>0,{noi_f}/{ds_f},0)'
        ws.cell(row=row, column=col).number_format = NF_PCT2
    row += 1

    # Cash-on-Cash
    ws.cell(row=row, column=2).value = 'Cash-on-Cash (Ann.)'
    ws.cell(row=row, column=2).font = F_LABEL
    fill = P_ALT if row % 2 == 0 else P_WHITE
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill
    for i, m in enumerate(months):
        col = first_data_col + i
        cfads_f = sumifs_month('CASH FLOW AFTER DEBT SERVICE', m)
        ws.cell(row=row, column=col).value = f'={cfads_f}*12/{cfg["total_equity"]}'
        ws.cell(row=row, column=col).number_format = NF_PCT2
    row += 1

    hide_beyond(ws, max(row + 2, 56), max_col)
    return ws
