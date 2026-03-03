# src/sheets/rr_input.py — Rent Roll raw data input
import csv
from datetime import datetime
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from design import (apply_title, apply_hdr, apply_subtotal, apply_alt_rows,
                     set_col_widths, hide_beyond,
                     F_BODY, F_BOLD, F_LABEL, F_MUTED,
                     P_ALT, P_WHITE, NF_DOLLAR, NF_NUM, NF_DATE)


def build(wb, cfg, rr_csv):
    ws = wb.create_sheet('RR Input')

    # Read rent roll CSV
    with open(rr_csv, 'r') as f:
        reader = csv.reader(f)
        header = next(reader)
        data = list(reader)

    months = header[1:]  # Column headers after 'Unit'
    num_months = len(months)
    max_col = 1 + num_months  # Col A=Unit, Cols B onwards = months

    apply_title(ws, max_col,
                'RENT ROLL INPUT',
                'Unit-level monthly rents. Source: rent_roll.csv. Zero indicates vacancy.')

    # Headers
    ws.cell(row=3, column=1).value = 'Unit'
    for i, m in enumerate(months):
        try:
            dt = datetime.strptime(m.strip(), '%Y-%m-%d')
            ws.cell(row=3, column=2 + i).value = dt
            ws.cell(row=3, column=2 + i).number_format = NF_DATE
        except ValueError:
            ws.cell(row=3, column=2 + i).value = m.strip()
    apply_hdr(ws, 3, max_col)

    from openpyxl.utils import get_column_letter
    widths = {'A': 10}
    for c in range(2, max_col + 1):
        widths[get_column_letter(c)] = 11
    set_col_widths(ws, widths)

    # Data rows
    row = 4
    for unit_data in data:
        unit = unit_data[0]
        fill = P_ALT if row % 2 == 0 else P_WHITE
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=1).value = unit
        ws.cell(row=row, column=1).font = F_LABEL
        for i, val in enumerate(unit_data[1:]):
            try:
                v = float(val) if val else 0
            except ValueError:
                v = 0
            ws.cell(row=row, column=2 + i).value = v
            ws.cell(row=row, column=2 + i).number_format = NF_DOLLAR
            ws.cell(row=row, column=2 + i).font = F_BODY
        row += 1

    # Total row
    total_row = row
    apply_subtotal(ws, total_row, max_col)
    ws.cell(row=total_row, column=1).value = 'TOTAL'
    ws.cell(row=total_row, column=1).font = F_BOLD
    for c in range(2, max_col + 1):
        cl = get_column_letter(c)
        ws.cell(row=total_row, column=c).value = f'=SUM({cl}4:{cl}{total_row - 1})'
        ws.cell(row=total_row, column=c).number_format = NF_DOLLAR
        ws.cell(row=total_row, column=c).font = F_BOLD

    row = total_row + 1

    # Occupied count
    row += 1
    ws.cell(row=row, column=1).value = 'Occupied'
    ws.cell(row=row, column=1).font = F_LABEL
    for c in range(2, max_col + 1):
        cl = get_column_letter(c)
        ws.cell(row=row, column=c).value = f'=COUNTIF({cl}4:{cl}{total_row - 1},">"&0)'
        ws.cell(row=row, column=c).number_format = NF_NUM
        ws.cell(row=row, column=c).font = F_BODY
        fill = P_ALT if row % 2 == 0 else P_WHITE
        ws.cell(row=row, column=c).fill = fill
    occ_row = row
    row += 1

    # Vacant count
    ws.cell(row=row, column=1).value = 'Vacant'
    ws.cell(row=row, column=1).font = F_LABEL
    for c in range(2, max_col + 1):
        cl = get_column_letter(c)
        ws.cell(row=row, column=c).value = f'=COUNTIF({cl}4:{cl}{total_row - 1},0)'
        ws.cell(row=row, column=c).number_format = NF_NUM
        ws.cell(row=row, column=c).font = F_BODY
        fill = P_ALT if row % 2 == 0 else P_WHITE
        ws.cell(row=row, column=c).fill = fill
    row += 1

    # Occupancy rate
    ws.cell(row=row, column=1).value = 'Occupancy %'
    ws.cell(row=row, column=1).font = F_LABEL
    units = cfg['property']['units']
    for c in range(2, max_col + 1):
        cl = get_column_letter(c)
        ws.cell(row=row, column=c).value = f'={cl}{occ_row}/{units}'
        ws.cell(row=row, column=c).number_format = '0.0%'
        ws.cell(row=row, column=c).font = F_BODY
        fill = P_ALT if row % 2 == 0 else P_WHITE
        ws.cell(row=row, column=c).fill = fill
    row += 1

    # Average rent
    ws.cell(row=row, column=1).value = 'Avg Rent'
    ws.cell(row=row, column=1).font = F_LABEL
    for c in range(2, max_col + 1):
        cl = get_column_letter(c)
        ws.cell(row=row, column=c).value = f'=IF({cl}{occ_row}>0,{cl}{total_row}/{cl}{occ_row},0)'
        ws.cell(row=row, column=c).number_format = NF_DOLLAR
        ws.cell(row=row, column=c).font = F_BODY
        fill = P_ALT if row % 2 == 0 else P_WHITE
        ws.cell(row=row, column=c).fill = fill
    row += 1

    hide_beyond(ws, max(row + 2, 131), max_col)
    return ws
