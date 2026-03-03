# src/sheets/t12_pl.py — Trailing 12-month P&L
import csv
from datetime import datetime
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from design import (apply_title, apply_hdr, apply_pl_formatting,
                     set_col_widths, hide_beyond,
                     F_BODY, F_BOLD, F_MUTED, F_GREEN,
                     P_ALT, P_WHITE, NF_DOLLAR, NF_PCT, NF_PCT2, NF_NUM, NF_DATE)


def build(wb, cfg):
    ws = wb.create_sheet('T12_PL')
    coa = cfg['coa']
    n = cfg['qpl_rows']

    # Determine T-12 months from data: read months from qPL_Fact
    # We need the last 12 months. Get them from pl_actuals months via config
    # Actually, read from qPL_Fact sheet
    qpl = wb['qPL_Fact']
    months = set()
    for r in range(2, n + 2):
        v = qpl.cell(row=r, column=1).value
        if v:
            if isinstance(v, datetime):
                months.add(v.strftime('%Y-%m-%d'))
            else:
                months.add(str(v))
    months = sorted(months)
    t12_months = months[-12:]

    # Layout: Col A=GL, Col B=Account, Cols 3-14=12 months, Col 15=T-12 Total,
    #         Col 16=$/Unit, Col 17=$/SF, Col 18=% of EGI
    first_data_col = 3
    num_t12 = len(t12_months)
    last_month_col = first_data_col + num_t12 - 1
    t12_col = last_month_col + 1
    unit_col = t12_col + 1
    sf_col = unit_col + 1
    pct_col = sf_col + 1
    max_col = pct_col

    apply_title(ws, max_col,
                'TRAILING 12-MONTH P&L',
                'Monthly detail for the trailing 12 months with totals and per-unit analysis.')

    # Headers
    ws.cell(row=3, column=1).value = 'GL'
    ws.cell(row=3, column=2).value = 'Account'
    for i, m in enumerate(t12_months):
        dt = datetime.strptime(m, '%Y-%m-%d')
        cell = ws.cell(row=3, column=first_data_col + i)
        cell.value = dt
        cell.number_format = NF_DATE
    ws.cell(row=3, column=t12_col).value = 'T-12 Total'
    ws.cell(row=3, column=unit_col).value = '$/Unit'
    ws.cell(row=3, column=sf_col).value = '$/SF'
    ws.cell(row=3, column=pct_col).value = '% of EGI'
    apply_hdr(ws, 3, max_col)

    from openpyxl.utils import get_column_letter
    widths = {'A': 6, 'B': 35}
    for c in range(first_data_col, last_month_col + 1):
        widths[get_column_letter(c)] = 12
    widths[get_column_letter(t12_col)] = 14
    widths[get_column_letter(unit_col)] = 12
    widths[get_column_letter(sf_col)] = 12
    widths[get_column_letter(pct_col)] = 10
    set_col_widths(ws, widths)

    sr = f'qPL_Fact!$D$2:$D${n}'
    ac = f'qPL_Fact!$C$2:$C${n}'
    dtr = f'qPL_Fact!$A$2:$A${n}'

    t12_start = datetime.strptime(t12_months[0], '%Y-%m-%d')
    t12_end = datetime.strptime(t12_months[-1], '%Y-%m-%d')
    date_gte = f'">="&DATE({t12_start.year},{t12_start.month},1)'
    date_lte = f'"<="&DATE({t12_end.year},{t12_end.month},1)'

    units = cfg['property']['units']
    rsf = cfg['property']['rentable_sf']
    egi_row = None

    row = 4
    for gl, account, rtype in coa:
        if rtype == 'spacer':
            row += 1
            continue

        if rtype == 'section':
            ws.cell(row=row, column=2).value = account
            row += 1
            continue

        ws.cell(row=row, column=1).value = gl
        ws.cell(row=row, column=2).value = account

        if account == 'EFFECTIVE GROSS INCOME (EGI)':
            egi_row = row

        if rtype in ('line', 'subtotal'):
            for i, m in enumerate(t12_months):
                dt = datetime.strptime(m, '%Y-%m-%d')
                col = first_data_col + i
                ws.cell(row=row, column=col).value = (
                    f'=SUMIFS({sr},{ac},"{account}",{dtr},DATE({dt.year},{dt.month},1))')
                ws.cell(row=row, column=col).number_format = NF_DOLLAR

            # T-12 total = sum of the 12 monthly cells
            first_cell = f'{get_column_letter(first_data_col)}{row}'
            last_cell = f'{get_column_letter(last_month_col)}{row}'
            ws.cell(row=row, column=t12_col).value = f'=SUM({first_cell}:{last_cell})'
            ws.cell(row=row, column=t12_col).number_format = NF_DOLLAR

            t12_ref = f'{get_column_letter(t12_col)}{row}'
            ws.cell(row=row, column=unit_col).value = f'={t12_ref}/{units}'
            ws.cell(row=row, column=unit_col).number_format = NF_DOLLAR

            ws.cell(row=row, column=sf_col).value = f'={t12_ref}/{rsf}'
            ws.cell(row=row, column=sf_col).number_format = NF_DOLLAR

            # % of EGI placeholder
            ws.cell(row=row, column=pct_col).number_format = NF_PCT

        elif rtype == 'metric':
            for i, m in enumerate(t12_months):
                dt = datetime.strptime(m, '%Y-%m-%d')
                col = first_data_col + i
                ws.cell(row=row, column=col).value = (
                    f'=SUMIFS({sr},{ac},"{account}",{dtr},DATE({dt.year},{dt.month},1))')
                ws.cell(row=row, column=col).number_format = NF_PCT2

            if account == 'DSCR':
                noi_f = f'SUMIFS({sr},{ac},"NET OPERATING INCOME (NOI)",{dtr},{date_gte},{dtr},{date_lte})'
                ds_f = f'SUMIFS({sr},{ac},"Total Debt Service",{dtr},{date_gte},{dtr},{date_lte})'
                ws.cell(row=row, column=t12_col).value = f'=IF({ds_f}<>0,{noi_f}/{ds_f},0)'
                ws.cell(row=row, column=t12_col).number_format = NF_PCT2
            elif account == 'Cap Rate':
                noi_f = f'SUMIFS({sr},{ac},"NET OPERATING INCOME (NOI)",{dtr},{date_gte},{dtr},{date_lte})'
                ws.cell(row=row, column=t12_col).value = f'={noi_f}/{cfg["purchase_price"]}'
                ws.cell(row=row, column=t12_col).number_format = NF_PCT2
            elif account == 'Expense Ratio':
                egi_f = f'SUMIFS({sr},{ac},"EFFECTIVE GROSS INCOME (EGI)",{dtr},{date_gte},{dtr},{date_lte})'
                opex_f = f'SUMIFS({sr},{ac},"Total Operating Expenses",{dtr},{date_gte},{dtr},{date_lte})'
                ws.cell(row=row, column=t12_col).value = f'=IF({egi_f}<>0,{opex_f}/{egi_f},0)'
                ws.cell(row=row, column=t12_col).number_format = NF_PCT2
            elif 'Cash-on-Cash' in account:
                cfads_f = f'SUMIFS({sr},{ac},"CASH FLOW AFTER DEBT SERVICE",{dtr},{date_gte},{dtr},{date_lte})'
                ws.cell(row=row, column=t12_col).value = f'={cfads_f}/{cfg["total_equity"]}'
                ws.cell(row=row, column=t12_col).number_format = NF_PCT2

        row += 1

    last_data_row = row - 1

    # Fill in % of EGI formulas
    if egi_row:
        egi_t12 = f'${get_column_letter(t12_col)}${egi_row}'
        for r in range(4, last_data_row + 1):
            cell = ws.cell(row=r, column=pct_col)
            t12_ref = f'{get_column_letter(t12_col)}{r}'
            acct = ws.cell(row=r, column=2).value
            if acct and acct not in ('REVENUE', 'OTHER INCOME', 'OPERATING EXPENSES',
                                      'DEBT SERVICE', 'CAPITAL EXPENDITURES (Below the Line)',
                                      'KEY METRICS'):
                rtype_check = None
                for _, a, rt in coa:
                    if a == acct:
                        rtype_check = rt
                        break
                if rtype_check in ('line', 'subtotal'):
                    cell.value = f'=IF({egi_t12}<>0,{t12_ref}/{egi_t12},0)'

    apply_pl_formatting(ws, max_col, data_start=4, data_end=last_data_row)
    hide_beyond(ws, max(last_data_row + 2, 91), max_col)
    return ws
