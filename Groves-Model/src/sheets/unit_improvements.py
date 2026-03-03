# src/sheets/unit_improvements.py — Renovation tracker
import csv
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from design import (apply_title, apply_hdr, apply_subtotal, apply_section,
                     set_col_widths, hide_beyond,
                     F_BODY, F_BOLD, F_LABEL, F_MUTED, F_GREEN,
                     P_ALT, P_WHITE, P_GREEN_LT,
                     NF_DOLLAR, NF_NUM, NF_PCT)


def build(wb, cfg, csv_path):
    ws = wb.create_sheet('Unit Improvements')

    # Read CSV
    with open(csv_path, 'r') as f:
        reader = csv.reader(f)
        header = next(reader)
        data = list(reader)

    num_cols = len(header)
    max_col = min(num_cols, 20)  # cap at 20 columns

    apply_title(ws, max_col,
                'UNIT IMPROVEMENTS',
                'Renovation tracker per unit. Source: unit_improvements.csv.')

    # Headers
    for c, h in enumerate(header[:max_col], 1):
        ws.cell(row=3, column=c).value = h
    apply_hdr(ws, 3, max_col)

    from openpyxl.utils import get_column_letter
    widths = {'A': 10, 'B': 8}
    for c in range(3, max_col + 1):
        widths[get_column_letter(c)] = 11
    # Wider columns for key fields
    widths['C'] = 12  # Condition
    widths['D'] = 11  # OrigRent
    widths['E'] = 11  # CurrRent
    widths['F'] = 11  # MktRent
    widths['G'] = 8   # LTL
    set_col_widths(ws, widths)

    # Money columns (OrigRent=D/4, CurrRent=E/5, MktRent=F/6, LTL=G/7)
    dollar_cols = {3, 4, 5, 6}  # 0-indexed from header: OrigRent(3), CurrRent(4), MktRent(5), LTL(6)

    row = 4
    for unit_data in data:
        fill = P_ALT if row % 2 == 0 else P_WHITE
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).fill = fill

        for c, val in enumerate(unit_data[:max_col], 1):
            cell = ws.cell(row=row, column=c)
            # First column is unit number
            if c == 1:
                cell.value = val
                cell.font = F_LABEL
            elif c <= 3:
                cell.value = val
                cell.font = F_BODY
            elif c <= 7:
                # Numeric columns (OrigRent, CurrRent, MktRent, LTL)
                try:
                    cell.value = float(val) if val else 0
                    cell.number_format = NF_DOLLAR if c <= 6 else NF_NUM
                except ValueError:
                    cell.value = val
                cell.font = F_BODY
            else:
                # Checkbox columns
                cell.value = val
                cell.font = F_MUTED
        row += 1

    # Summary section
    last_data_row = row - 1
    row += 1

    apply_section(ws, row, max_col, accent=True)
    ws.cell(row=row, column=2).value = 'SUMMARY'
    row += 1

    # Count renovated
    ws.cell(row=row, column=1).value = 'Total Units Listed'
    ws.cell(row=row, column=1).font = F_LABEL
    ws.cell(row=row, column=4).value = f'=COUNTA(A4:A{last_data_row})'
    ws.cell(row=row, column=4).number_format = NF_NUM
    fill = P_ALT if row % 2 == 0 else P_WHITE
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill
    row += 1

    ws.cell(row=row, column=1).value = 'Avg Original Rent'
    ws.cell(row=row, column=1).font = F_LABEL
    ws.cell(row=row, column=4).value = f'=AVERAGE(D4:D{last_data_row})'
    ws.cell(row=row, column=4).number_format = NF_DOLLAR
    fill = P_ALT if row % 2 == 0 else P_WHITE
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill
    row += 1

    ws.cell(row=row, column=1).value = 'Avg Current Rent'
    ws.cell(row=row, column=1).font = F_LABEL
    ws.cell(row=row, column=5).value = f'=AVERAGE(E4:E{last_data_row})'
    ws.cell(row=row, column=5).number_format = NF_DOLLAR
    fill = P_ALT if row % 2 == 0 else P_WHITE
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill
    row += 1

    ws.cell(row=row, column=1).value = 'Avg Market Rent'
    ws.cell(row=row, column=1).font = F_LABEL
    ws.cell(row=row, column=6).value = f'=AVERAGE(F4:F{last_data_row})'
    ws.cell(row=row, column=6).number_format = NF_DOLLAR
    fill = P_ALT if row % 2 == 0 else P_WHITE
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill
    row += 1

    ws.cell(row=row, column=1).value = 'Avg Rent Increase'
    ws.cell(row=row, column=1).font = F_LABEL
    avg_curr = f'AVERAGE(E4:E{last_data_row})'
    avg_orig = f'AVERAGE(D4:D{last_data_row})'
    ws.cell(row=row, column=5).value = f'=IF({avg_orig}<>0,({avg_curr}-{avg_orig})/{avg_orig},0)'
    ws.cell(row=row, column=5).number_format = NF_PCT
    fill = P_ALT if row % 2 == 0 else P_WHITE
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill
    row += 1

    hide_beyond(ws, max(row + 2, 128), max_col)
    return ws
