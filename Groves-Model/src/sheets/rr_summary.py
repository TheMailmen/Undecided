# src/sheets/rr_summary.py — Rent Roll Summary
import csv
from datetime import datetime
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from design import (apply_title, apply_hdr, apply_section, apply_subtotal,
                     set_col_widths, hide_beyond,
                     F_BODY, F_BOLD, F_LABEL, F_GREEN, F_MUTED,
                     P_ALT, P_WHITE, P_GREEN_LT,
                     NF_DOLLAR, NF_PCT, NF_NUM, NF_DATE)


def build(wb, cfg):
    ws = wb.create_sheet('RR Summary')

    # Reference RR Input sheet
    rr = wb['RR Input']
    # Find months from RR Input header row (row 3)
    months = []
    col = 2
    while True:
        v = rr.cell(row=3, column=col).value
        if v is None:
            break
        months.append((col, v))
        col += 1

    num_months = len(months)
    # Layout: Col A=spacer, Col B=Label, Cols 3..3+num_months-1 = monthly data
    first_data_col = 3
    max_col = first_data_col + num_months - 1

    apply_title(ws, max_col,
                'RENT ROLL SUMMARY',
                'Aggregated occupancy, revenue, and average rents. References RR Input.')

    # Headers
    ws.cell(row=3, column=1).value = ''
    ws.cell(row=3, column=2).value = 'Metric'
    for i, (_, m) in enumerate(months):
        cell = ws.cell(row=3, column=first_data_col + i)
        cell.value = m
        if isinstance(m, datetime):
            cell.number_format = NF_DATE
    apply_hdr(ws, 3, max_col)

    from openpyxl.utils import get_column_letter
    widths = {'A': 4, 'B': 25}
    for c in range(first_data_col, max_col + 1):
        widths[get_column_letter(c)] = 12
    set_col_widths(ws, widths)

    # Find key rows in RR Input
    # Total row, Occupied row, etc. — scan for labels
    total_row_rr = None
    occ_row_rr = None
    vacant_row_rr = None
    occ_pct_row_rr = None
    avg_rent_row_rr = None

    for r in range(4, 140):
        v = rr.cell(row=r, column=1).value
        if v == 'TOTAL':
            total_row_rr = r
        elif v == 'Occupied':
            occ_row_rr = r
        elif v == 'Vacant':
            vacant_row_rr = r
        elif v == 'Occupancy %':
            occ_pct_row_rr = r
        elif v == 'Avg Rent':
            avg_rent_row_rr = r

    row = 4

    # ── OCCUPANCY ──
    apply_section(ws, row, max_col, accent=True)
    ws.cell(row=row, column=2).value = 'OCCUPANCY'
    row += 1

    # Total Rent
    ws.cell(row=row, column=2).value = 'Total Rent'
    ws.cell(row=row, column=2).font = F_BOLD
    for i, (rr_col, _) in enumerate(months):
        col = first_data_col + i
        rr_cl = get_column_letter(rr_col)
        ws.cell(row=row, column=col).value = f"='RR Input'!{rr_cl}{total_row_rr}"
        ws.cell(row=row, column=col).number_format = NF_DOLLAR
    apply_subtotal(ws, row, max_col)
    row += 1

    def summary_row(label, rr_row, nf, font=F_BODY):
        nonlocal row
        fill = P_ALT if row % 2 == 0 else P_WHITE
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font = F_LABEL
        if rr_row:
            for i, (rr_col, _) in enumerate(months):
                col = first_data_col + i
                rr_cl = get_column_letter(rr_col)
                ws.cell(row=row, column=col).value = f"='RR Input'!{rr_cl}{rr_row}"
                ws.cell(row=row, column=col).number_format = nf
                ws.cell(row=row, column=col).font = font
        row += 1

    summary_row('Units Occupied', occ_row_rr, NF_NUM)
    summary_row('Units Vacant', vacant_row_rr, NF_NUM)
    summary_row('Occupancy Rate', occ_pct_row_rr, NF_PCT)
    summary_row('Average Rent', avg_rent_row_rr, NF_DOLLAR)

    row += 1

    # ── UNIT MIX ANALYSIS ──
    apply_section(ws, row, max_col, accent=True)
    ws.cell(row=row, column=2).value = 'UNIT MIX'
    row += 1

    umix = cfg['unit_mix']
    for utype, info in umix.items():
        fill = P_ALT if row % 2 == 0 else P_WHITE
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=2).value = f'{utype} — {info["count"]} units @ {info["sf"]} SF'
        ws.cell(row=row, column=2).font = F_LABEL
        ws.cell(row=row, column=first_data_col).value = f'Market: ${info["market_rent"]}'
        ws.cell(row=row, column=first_data_col).font = F_MUTED
        row += 1

    row += 1

    # ── MONTHLY TRENDS ──
    apply_section(ws, row, max_col, accent=True)
    ws.cell(row=row, column=2).value = 'MONTHLY RENT TRENDS'
    row += 1

    # Month-over-month change in total rent
    ws.cell(row=row, column=2).value = 'MoM Change ($)'
    ws.cell(row=row, column=2).font = F_LABEL
    fill = P_ALT if row % 2 == 0 else P_WHITE
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill
    for i, (rr_col, _) in enumerate(months):
        col = first_data_col + i
        rr_cl = get_column_letter(rr_col)
        if i == 0:
            ws.cell(row=row, column=col).value = 0
        else:
            prev_rr_cl = get_column_letter(months[i-1][0])
            ws.cell(row=row, column=col).value = (
                f"='RR Input'!{rr_cl}{total_row_rr}-'RR Input'!{prev_rr_cl}{total_row_rr}")
        ws.cell(row=row, column=col).number_format = NF_DOLLAR
    row += 1

    # MoM change %
    ws.cell(row=row, column=2).value = 'MoM Change (%)'
    ws.cell(row=row, column=2).font = F_LABEL
    fill = P_ALT if row % 2 == 0 else P_WHITE
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill
    for i, (rr_col, _) in enumerate(months):
        col = first_data_col + i
        if i == 0:
            ws.cell(row=row, column=col).value = 0
        else:
            prev_cl = get_column_letter(col - 1)
            rr_prev_cl = get_column_letter(months[i-1][0])
            this_cl = get_column_letter(col)
            # Use the row above (MoM Change $) / prior month total
            mom_row = row - 1
            ws.cell(row=row, column=col).value = (
                f"=IF('RR Input'!{rr_prev_cl}{total_row_rr}<>0,"
                f"{this_cl}{mom_row}/'RR Input'!{rr_prev_cl}{total_row_rr},0)")
        ws.cell(row=row, column=col).number_format = NF_PCT
    row += 1

    row += 1

    # ── REVENUE POTENTIAL ──
    apply_section(ws, row, max_col, accent=True)
    ws.cell(row=row, column=2).value = 'REVENUE POTENTIAL'
    row += 1

    total_units = cfg['property']['units']
    total_market = sum(info['count'] * info['market_rent'] for info in umix.values())

    fill = P_ALT if row % 2 == 0 else P_WHITE
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill
    ws.cell(row=row, column=2).value = 'Gross Potential (all at market)'
    ws.cell(row=row, column=2).font = F_LABEL
    ws.cell(row=row, column=first_data_col).value = total_market
    ws.cell(row=row, column=first_data_col).number_format = NF_DOLLAR
    row += 1

    fill = P_ALT if row % 2 == 0 else P_WHITE
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill
    ws.cell(row=row, column=2).value = 'Annual Market Revenue'
    ws.cell(row=row, column=2).font = F_LABEL
    ws.cell(row=row, column=first_data_col).value = total_market * 12
    ws.cell(row=row, column=first_data_col).number_format = NF_DOLLAR
    row += 1

    hide_beyond(ws, max(row + 2, 37), max_col)
    return ws
