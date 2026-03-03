# src/sheets/trailing.py — Trailing analysis (T-1, T-3, T-6, T-12)
import csv
from datetime import datetime
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from design import (apply_title, apply_hdr, apply_pl_formatting,
                     set_col_widths, hide_beyond,
                     F_BODY, F_BOLD, F_MUTED, F_GREEN,
                     P_ALT, P_WHITE, NF_DOLLAR, NF_PCT, NF_PCT2, NF_NUM, NF_DATE)


def build(wb, cfg):
    ws = wb.create_sheet('Trailing Analysis')
    coa = cfg['coa']
    n = cfg['qpl_rows']

    # Get months from qPL_Fact
    qpl = wb['qPL_Fact']
    months = set()
    for r in range(2, n + 2):
        v = qpl.cell(row=r, column=1).value
        if v:
            if isinstance(v, datetime):
                months.add(v.strftime('%Y-%m-%d'))
            else:
                months.add(str(v))
    months = sorted(months)

    # Define trailing windows
    windows = [
        ('T-1', months[-1:]),
        ('T-3', months[-3:]),
        ('T-6', months[-6:]),
        ('T-12', months[-12:]),
    ]

    # Layout: Col A=GL, Col B=Account,
    # For each window: monthly columns + total = multiple sections
    # Simpler approach: just show summary columns for each window
    # Col 3-6: T-1 Actual, T-3 Total, T-6 Total, T-12 Total
    # Col 7-10: T-1 Ann., T-3 Ann., T-6 Ann., T-12 (actual)
    # Col 11-14: T-1 $/Unit, T-3 $/Unit, T-6 $/Unit, T-12 $/Unit
    # Col 15-18: T-1 $/SF, T-3 $/SF, T-6 $/SF, T-12 $/SF
    # Col 19-20: Var T-3 vs T-12, Var T-6 vs T-12

    # Actually let me keep it simpler and more useful:
    # Cols: GL, Account, T-1 Month, T-3 Total, T-3 Ann., T-6 Total, T-6 Ann.,
    #        T-12 Total, $/Unit, $/SF, T-1 vs T-12 Var

    max_col = 20
    first_data_col = 3

    apply_title(ws, max_col,
                'TRAILING ANALYSIS',
                'Rolling performance windows: T-1, T-3, T-6, and T-12 with annualized comparisons.')

    # Headers
    headers = ['GL', 'Account',
               'T-1', 'T-1 Ann.',
               'T-3 Total', 'T-3 Ann.',
               'T-6 Total', 'T-6 Ann.',
               'T-12 Total', 'T-12 $/Unit', 'T-12 $/SF',
               '', # spacer
               'T-1 Ann. Var', 'T-3 Ann. Var', 'T-6 Ann. Var',
               '', # spacer
               'T-1 Margin', 'T-3 Margin', 'T-6 Margin', 'T-12 Margin']

    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c).value = h
    apply_hdr(ws, 3, max_col)

    from openpyxl.utils import get_column_letter
    widths = {'A': 6, 'B': 35}
    for c in range(3, max_col + 1):
        widths[get_column_letter(c)] = 13
    widths[get_column_letter(12)] = 2  # spacer
    widths[get_column_letter(16)] = 2  # spacer
    set_col_widths(ws, widths)

    sr = f'qPL_Fact!$D$2:$D${n}'
    ac = f'qPL_Fact!$C$2:$C${n}'
    dtr = f'qPL_Fact!$A$2:$A${n}'

    units = cfg['property']['units']
    rsf = cfg['property']['rentable_sf']

    def date_range(window_months):
        s = datetime.strptime(window_months[0], '%Y-%m-%d')
        e = datetime.strptime(window_months[-1], '%Y-%m-%d')
        return (f'">="&DATE({s.year},{s.month},1)',
                f'"<="&DATE({e.year},{e.month},1)')

    def sumifs_range(account, window_months):
        if len(window_months) == 1:
            dt = datetime.strptime(window_months[0], '%Y-%m-%d')
            return f'SUMIFS({sr},{ac},"{account}",{dtr},DATE({dt.year},{dt.month},1))'
        gte, lte = date_range(window_months)
        return f'SUMIFS({sr},{ac},"{account}",{dtr},{gte},{dtr},{lte})'

    egi_row = None
    row = 4

    for gl, account, rtype in coa:
        if rtype == 'spacer':
            row += 1
            continue

        if rtype == 'section':
            ws.cell(row=row, column=2).value = account
            row += 1
            continue

        ws.cell(row=row, column=1).value = gl
        ws.cell(row=row, column=2).value = account

        if account == 'EFFECTIVE GROSS INCOME (EGI)':
            egi_row = row

        if rtype in ('line', 'subtotal'):
            # T-1 (col 3)
            w1 = windows[0][1]
            ws.cell(row=row, column=3).value = f'={sumifs_range(account, w1)}'
            ws.cell(row=row, column=3).number_format = NF_DOLLAR
            # T-1 Ann. (col 4) = T-1 * 12
            ws.cell(row=row, column=4).value = f'=C{row}*12'
            ws.cell(row=row, column=4).number_format = NF_DOLLAR

            # T-3 (col 5)
            w3 = windows[1][1]
            ws.cell(row=row, column=5).value = f'={sumifs_range(account, w3)}'
            ws.cell(row=row, column=5).number_format = NF_DOLLAR
            # T-3 Ann. (col 6) = T-3 * 4
            ws.cell(row=row, column=6).value = f'=E{row}*4'
            ws.cell(row=row, column=6).number_format = NF_DOLLAR

            # T-6 (col 7)
            w6 = windows[2][1]
            ws.cell(row=row, column=7).value = f'={sumifs_range(account, w6)}'
            ws.cell(row=row, column=7).number_format = NF_DOLLAR
            # T-6 Ann. (col 8) = T-6 * 2
            ws.cell(row=row, column=8).value = f'=G{row}*2'
            ws.cell(row=row, column=8).number_format = NF_DOLLAR

            # T-12 (col 9)
            w12 = windows[3][1]
            ws.cell(row=row, column=9).value = f'={sumifs_range(account, w12)}'
            ws.cell(row=row, column=9).number_format = NF_DOLLAR

            # T-12 $/Unit (col 10)
            ws.cell(row=row, column=10).value = f'=I{row}/{units}'
            ws.cell(row=row, column=10).number_format = NF_DOLLAR

            # T-12 $/SF (col 11)
            ws.cell(row=row, column=11).value = f'=I{row}/{rsf}'
            ws.cell(row=row, column=11).number_format = NF_DOLLAR

            # Col 12 = spacer

            # Variance cols: annualized vs T-12
            # T-1 Ann. Var (col 13) = (T-1 Ann - T-12) / T-12
            ws.cell(row=row, column=13).value = f'=IF(I{row}<>0,(D{row}-I{row})/ABS(I{row}),0)'
            ws.cell(row=row, column=13).number_format = NF_PCT
            # T-3 Ann. Var (col 14)
            ws.cell(row=row, column=14).value = f'=IF(I{row}<>0,(F{row}-I{row})/ABS(I{row}),0)'
            ws.cell(row=row, column=14).number_format = NF_PCT
            # T-6 Ann. Var (col 15)
            ws.cell(row=row, column=15).value = f'=IF(I{row}<>0,(H{row}-I{row})/ABS(I{row}),0)'
            ws.cell(row=row, column=15).number_format = NF_PCT

            # Col 16 = spacer

            # Margin cols (% of EGI for each window)
            # Will fill after we know EGI row

        elif rtype == 'metric':
            # Metrics don't sum — show the latest value
            w1 = windows[0][1]
            ws.cell(row=row, column=3).value = f'={sumifs_range(account, w1)}'
            ws.cell(row=row, column=3).number_format = NF_PCT2

            for cidx in (4, 5, 6, 7, 8):
                ws.cell(row=row, column=cidx).number_format = NF_PCT2

            # T-12 metrics
            if account == 'DSCR':
                noi_f = sumifs_range('NET OPERATING INCOME (NOI)', windows[3][1])
                ds_f = sumifs_range('Total Debt Service', windows[3][1])
                ws.cell(row=row, column=9).value = f'=IF({ds_f}<>0,{noi_f}/{ds_f},0)'
            elif account == 'Cap Rate':
                noi_f = sumifs_range('NET OPERATING INCOME (NOI)', windows[3][1])
                ws.cell(row=row, column=9).value = f'={noi_f}/{cfg["purchase_price"]}'
            elif account == 'Expense Ratio':
                egi_f = sumifs_range('EFFECTIVE GROSS INCOME (EGI)', windows[3][1])
                opex_f = sumifs_range('Total Operating Expenses', windows[3][1])
                ws.cell(row=row, column=9).value = f'=IF({egi_f}<>0,{opex_f}/{egi_f},0)'
            elif 'Cash-on-Cash' in account:
                cfads_f = sumifs_range('CASH FLOW AFTER DEBT SERVICE', windows[3][1])
                ws.cell(row=row, column=9).value = f'={cfads_f}/{cfg["total_equity"]}'
            ws.cell(row=row, column=9).number_format = NF_PCT2

        row += 1

    last_data_row = row - 1

    # Fill margin columns (17-20) — % of EGI for each window
    if egi_row:
        for r in range(4, last_data_row + 1):
            acct = ws.cell(row=r, column=2).value
            if not acct:
                continue
            rtype_check = None
            for _, a, rt in coa:
                if a == acct:
                    rtype_check = rt
                    break
            if rtype_check in ('line', 'subtotal'):
                # T-1 margin
                ws.cell(row=r, column=17).value = f'=IF(C${egi_row}<>0,C{r}/C${egi_row},0)'
                ws.cell(row=r, column=17).number_format = NF_PCT
                # T-3 margin
                ws.cell(row=r, column=18).value = f'=IF(E${egi_row}<>0,E{r}/E${egi_row},0)'
                ws.cell(row=r, column=18).number_format = NF_PCT
                # T-6 margin
                ws.cell(row=r, column=19).value = f'=IF(G${egi_row}<>0,G{r}/G${egi_row},0)'
                ws.cell(row=r, column=19).number_format = NF_PCT
                # T-12 margin
                ws.cell(row=r, column=20).value = f'=IF(I${egi_row}<>0,I{r}/I${egi_row},0)'
                ws.cell(row=r, column=20).number_format = NF_PCT

    apply_pl_formatting(ws, max_col, data_start=4, data_end=last_data_row)
    hide_beyond(ws, max(last_data_row + 2, 93), max_col)
    return ws
