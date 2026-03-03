# src/sheets/full_pl.py — Full monthly P&L
import csv
from datetime import datetime
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from design import (apply_title, apply_hdr, apply_pl_formatting,
                     set_col_widths, hide_beyond,
                     F_BODY, F_BOLD, F_MUTED, F_GREEN,
                     P_ALT, P_WHITE, NF_DOLLAR, NF_PCT, NF_PCT2, NF_NUM, NF_DATE)


def build(wb, cfg, pl_csv):
    ws = wb.create_sheet('Full P&L')
    coa = cfg['coa']
    n = cfg['qpl_rows']

    # Read months from CSV
    months = set()
    with open(pl_csv, 'r') as f:
        for row in csv.DictReader(f):
            months.add(row['Month'].strip())
    months = sorted(months)
    num_months = len(months)

    # Layout:
    # Col A(1)=GL, Col B(2)=Account
    # Cols 3..3+num_months-1 = monthly data
    # Col after months+1 = spacer
    # Then: T-12 Total, $/Unit, $/SF, % of EGI
    first_data_col = 3
    last_month_col = first_data_col + num_months - 1
    spacer_col = last_month_col + 1
    t12_col = spacer_col + 1
    unit_col = t12_col + 1
    sf_col = unit_col + 1
    pct_col = sf_col + 1
    max_col = pct_col

    apply_title(ws, max_col,
                'FULL P&L — THE GROVES APARTMENTS',
                'Monthly actuals from source data. All values are SUMIFS formulas against qPL_Fact.')

    # Headers
    ws.cell(row=3, column=1).value = 'GL'
    ws.cell(row=3, column=2).value = 'Account'
    for i, m in enumerate(months):
        dt = datetime.strptime(m, '%Y-%m-%d')
        cell = ws.cell(row=3, column=first_data_col + i)
        cell.value = dt
        cell.number_format = NF_DATE
    ws.cell(row=3, column=t12_col).value = 'T-12 Total'
    ws.cell(row=3, column=unit_col).value = '$/Unit'
    ws.cell(row=3, column=sf_col).value = '$/SF'
    ws.cell(row=3, column=pct_col).value = '% of EGI'
    apply_hdr(ws, 3, max_col)

    # Column widths
    widths = {'A': 6, 'B': 35}
    from openpyxl.utils import get_column_letter
    for c in range(first_data_col, last_month_col + 1):
        widths[get_column_letter(c)] = 12
    widths[get_column_letter(spacer_col)] = 2
    widths[get_column_letter(t12_col)] = 14
    widths[get_column_letter(unit_col)] = 12
    widths[get_column_letter(sf_col)] = 12
    widths[get_column_letter(pct_col)] = 10
    set_col_widths(ws, widths)

    # SUMIFS refs
    sr = f'qPL_Fact!$D$2:$D${n}'
    ac = f'qPL_Fact!$C$2:$C${n}'
    dtr = f'qPL_Fact!$A$2:$A${n}'

    # T-12 date range (last 12 months)
    t12_months = months[-12:]
    t12_start = datetime.strptime(t12_months[0], '%Y-%m-%d')
    t12_end = datetime.strptime(t12_months[-1], '%Y-%m-%d')
    date_gte = f'">="&DATE({t12_start.year},{t12_start.month},1)'
    date_lte = f'"<="&DATE({t12_end.year},{t12_end.month},1)'

    units = cfg['property']['units']
    rsf = cfg['property']['rentable_sf']

    # Track EGI row for % of EGI calculation
    egi_row = None

    row = 4
    for gl, account, rtype in coa:
        if rtype == 'spacer':
            row += 1
            continue

        if rtype == 'section':
            ws.cell(row=row, column=2).value = account
            # Sections have no data
            row += 1
            continue

        ws.cell(row=row, column=1).value = gl
        ws.cell(row=row, column=2).value = account

        if account == 'EFFECTIVE GROSS INCOME (EGI)':
            egi_row = row

        is_metric = rtype == 'metric'

        if rtype in ('line', 'subtotal'):
            # Monthly SUMIFS
            for i, m in enumerate(months):
                dt = datetime.strptime(m, '%Y-%m-%d')
                col = first_data_col + i
                formula = f'=SUMIFS({sr},{ac},"{account}",{dtr},DATE({dt.year},{dt.month},1))'
                ws.cell(row=row, column=col).value = formula
                ws.cell(row=row, column=col).number_format = NF_DOLLAR

            # T-12 total
            t12_formula = f'=SUMIFS({sr},{ac},"{account}",{dtr},{date_gte},{dtr},{date_lte})'
            ws.cell(row=row, column=t12_col).value = t12_formula
            ws.cell(row=row, column=t12_col).number_format = NF_DOLLAR

            # $/Unit
            t12_cell = f'{get_column_letter(t12_col)}{row}'
            ws.cell(row=row, column=unit_col).value = f'={t12_cell}/{units}'
            ws.cell(row=row, column=unit_col).number_format = NF_DOLLAR

            # $/SF
            ws.cell(row=row, column=sf_col).value = f'={t12_cell}/{rsf}'
            ws.cell(row=row, column=sf_col).number_format = NF_DOLLAR

            # % of EGI (placeholder — will update once we know EGI row)
            ws.cell(row=row, column=pct_col).value = f'=IF(${get_column_letter(t12_col)}$__EGI__<>0,{t12_cell}/${get_column_letter(t12_col)}$__EGI__,0)'
            ws.cell(row=row, column=pct_col).number_format = NF_PCT

        elif is_metric:
            # Metrics: DSCR, Cap Rate, Expense Ratio, Cash-on-Cash
            for i, m in enumerate(months):
                dt = datetime.strptime(m, '%Y-%m-%d')
                col = first_data_col + i
                formula = f'=SUMIFS({sr},{ac},"{account}",{dtr},DATE({dt.year},{dt.month},1))'
                ws.cell(row=row, column=col).value = formula
                if account in ('DSCR', 'Cap Rate', 'Expense Ratio',
                               'Cash-on-Cash (CFADS)', 'Cash-on-Cash (Ann.)'):
                    ws.cell(row=row, column=col).number_format = NF_PCT2
                else:
                    ws.cell(row=row, column=col).number_format = NF_DOLLAR

            # T-12 metric — cannot sum; compute from T-12 totals
            if account == 'DSCR':
                noi_f = f'SUMIFS({sr},{ac},"NET OPERATING INCOME (NOI)",{dtr},{date_gte},{dtr},{date_lte})'
                ds_f = f'SUMIFS({sr},{ac},"Total Debt Service",{dtr},{date_gte},{dtr},{date_lte})'
                ws.cell(row=row, column=t12_col).value = f'=IF({ds_f}<>0,{noi_f}/{ds_f},0)'
                ws.cell(row=row, column=t12_col).number_format = NF_PCT2
            elif account == 'Cap Rate':
                noi_f = f'SUMIFS({sr},{ac},"NET OPERATING INCOME (NOI)",{dtr},{date_gte},{dtr},{date_lte})'
                ws.cell(row=row, column=t12_col).value = f'={noi_f}/{cfg["purchase_price"]}'
                ws.cell(row=row, column=t12_col).number_format = NF_PCT2
            elif account == 'Expense Ratio':
                egi_f = f'SUMIFS({sr},{ac},"EFFECTIVE GROSS INCOME (EGI)",{dtr},{date_gte},{dtr},{date_lte})'
                opex_f = f'SUMIFS({sr},{ac},"Total Operating Expenses",{dtr},{date_gte},{dtr},{date_lte})'
                ws.cell(row=row, column=t12_col).value = f'=IF({egi_f}<>0,{opex_f}/{egi_f},0)'
                ws.cell(row=row, column=t12_col).number_format = NF_PCT2
            elif 'Cash-on-Cash' in account:
                cfads_f = f'SUMIFS({sr},{ac},"CASH FLOW AFTER DEBT SERVICE",{dtr},{date_gte},{dtr},{date_lte})'
                ws.cell(row=row, column=t12_col).value = f'={cfads_f}/{cfg["total_equity"]}'
                ws.cell(row=row, column=t12_col).number_format = NF_PCT2

        row += 1

    last_data_row = row - 1

    # Fix EGI references in % of EGI column
    if egi_row:
        for r in range(4, last_data_row + 1):
            cell = ws.cell(row=r, column=pct_col)
            if cell.value and isinstance(cell.value, str) and '__EGI__' in cell.value:
                cell.value = cell.value.replace('__EGI__', str(egi_row))

    # Apply P&L formatting
    apply_pl_formatting(ws, max_col, data_start=4, data_end=last_data_row)

    hide_beyond(ws, max(last_data_row + 2, 112), max_col)
    return ws
