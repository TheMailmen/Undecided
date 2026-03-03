# src/design.py — Single source of truth for all formatting
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# COLORS
# ============================================================
C_TITLE    = '0D1B2A'
C_SUB      = '1A2332'
C_HDR      = '0D1B2A'
C_ACCENT   = '1B4F72'
C_GREEN    = '1E8449'
C_RED      = 'C00000'
C_GOLD     = 'D4AC0D'
C_SUBTOTAL = 'E8EAED'
C_ALT      = 'F7F9FC'
C_WHITE    = 'FFFFFF'
C_INPUT_BG = 'FFF9C4'
C_NOTE_BG  = 'E8F5E9'
C_WARN_BG  = 'FDEDEC'
C_BODY     = '000000'
C_LABEL    = '2C3E50'
C_MUTED    = '7F8C8D'
C_BLUE     = '0000FF'
C_CROSS    = '548235'

# ============================================================
# FONTS
# ============================================================
F_TITLE   = Font(name='Calibri', size=14, bold=True, color=C_WHITE)
F_SUB     = Font(name='Calibri', size=9, color='AEB6BF')
F_HDR     = Font(name='Calibri', size=10, bold=True, color=C_WHITE)
F_SECTION = Font(name='Calibri', size=10, bold=True, color=C_WHITE)
F_LABEL   = Font(name='Calibri', size=10, bold=True, color=C_LABEL)
F_BODY    = Font(name='Calibri', size=10, color=C_BODY)
F_BOLD    = Font(name='Calibri', size=10, bold=True, color=C_BODY)
F_INPUT   = Font(name='Calibri', size=10, bold=True, color=C_BLUE)
F_MUTED   = Font(name='Calibri', size=9, color=C_MUTED)
F_GREEN   = Font(name='Calibri', size=10, bold=True, color=C_GREEN)
F_RED     = Font(name='Calibri', size=10, bold=True, color=C_RED)
F_CROSS   = Font(name='Calibri', size=10, color=C_CROSS)
F_NOTE    = Font(name='Calibri', size=9, italic=True, color=C_GREEN)
F_GOLD    = Font(name='Calibri', size=10, bold=True, color=C_GOLD)

# ============================================================
# FILLS
# ============================================================
P_TITLE    = PatternFill('solid', fgColor=C_TITLE)
P_SUB      = PatternFill('solid', fgColor=C_SUB)
P_HDR      = PatternFill('solid', fgColor=C_HDR)
P_ACCENT   = PatternFill('solid', fgColor=C_ACCENT)
P_SUBTOTAL = PatternFill('solid', fgColor=C_SUBTOTAL)
P_ALT      = PatternFill('solid', fgColor=C_ALT)
P_WHITE    = PatternFill('solid', fgColor=C_WHITE)
P_INPUT    = PatternFill('solid', fgColor=C_INPUT_BG)
P_GREEN_LT = PatternFill('solid', fgColor=C_NOTE_BG)
P_RED_LT   = PatternFill('solid', fgColor=C_WARN_BG)
P_NONE     = PatternFill()

# ============================================================
# BORDERS
# ============================================================
B_THIN    = Border(bottom=Side(style='thin', color='D9D9D9'))
B_SUB     = Border(top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
B_SECTION = Border(bottom=Side(style='medium', color=C_HDR))
B_NONE    = Border()

# ============================================================
# ALIGNMENT
# ============================================================
A_CENTER = Alignment(horizontal='center', vertical='center')
A_LEFT   = Alignment(horizontal='left', vertical='center')
A_RIGHT  = Alignment(horizontal='right', vertical='center')

# ============================================================
# NUMBER FORMATS
# ============================================================
NF_DOLLAR = '$#,##0;($#,##0);"-"'
NF_NUM    = '#,##0;(#,##0);"-"'
NF_PCT    = '0.0%'
NF_PCT2   = '0.00%'
NF_DATE   = 'MMM-YY'
NF_DEC2   = '#,##0.00'

# ============================================================
# P&L CATEGORY SETS (used across multiple sheets)
# ============================================================
SECTIONS = {
    'REVENUE', 'OTHER INCOME', 'OPERATING EXPENSES', 'DEBT SERVICE',
    'CAPITAL EXPENDITURES (Below the Line)', 'KEY METRICS',
    'ESCROW & RESERVE ACCOUNTS', 'Operating Account',
}

SUBTOTALS = {
    'Effective Rental Income', 'Total Other Income',
    'EFFECTIVE GROSS INCOME (EGI)', 'Total Operating Expenses',
    'NET OPERATING INCOME (NOI)', 'Total Debt Service',
    'CASH FLOW AFTER DEBT SERVICE', 'Total Capital Expenditures',
    'NET CASH FLOW',
}

GREEN_ROWS = {
    'NET OPERATING INCOME (NOI)', 'CASH FLOW AFTER DEBT SERVICE',
    'NET CASH FLOW', 'FREE CASH FOR DISTRIBUTION',
}

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def apply_title(ws, max_col, title, subtitle):
    """Apply title bar (row 1) and subtitle bar (row 2)."""
    for c in range(1, max_col + 1):
        ws.cell(row=1, column=c).fill = P_TITLE
        ws.cell(row=2, column=c).fill = P_SUB
    ws.cell(row=1, column=1).value = title
    ws.cell(row=1, column=1).font = F_TITLE
    ws.cell(row=2, column=1).value = subtitle
    ws.cell(row=2, column=1).font = F_SUB
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18


def apply_hdr(ws, row, max_col):
    """Apply dark column header row."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = F_HDR
        cell.fill = P_HDR
        cell.alignment = A_CENTER
    ws.row_dimensions[row].height = 20


def apply_section(ws, row, max_col, accent=False):
    """Apply section divider bar."""
    fill = P_ACCENT if accent else P_HDR
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = F_SECTION
    ws.row_dimensions[row].height = 20


def apply_subtotal(ws, row, max_col, green=False):
    """Apply subtotal row formatting."""
    fill = P_GREEN_LT if green else P_SUBTOTAL
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.border = B_SUB
        if c > 2:
            cell.font = F_BOLD


def apply_alt_rows(ws, start_row, end_row, max_col):
    """Apply alternating white / light gray row fills."""
    for row in range(start_row, end_row + 1):
        fill = P_ALT if (row - start_row) % 2 == 0 else P_WHITE
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).fill = fill


def apply_pl_formatting(ws, max_col, data_start=4, data_end=None):
    """Apply P&L-style formatting to any sheet with account names in col B.
    Handles sections, subtotals, green rows, GL codes, alternating fills."""
    if data_end is None:
        data_end = ws.max_row
    for row in range(data_start, data_end + 1):
        acct = ws.cell(row=row, column=2).value
        gl = ws.cell(row=row, column=1).value
        if acct in SECTIONS:
            apply_section(ws, row, max_col, accent=True)
        elif acct in SUBTOTALS:
            is_green = acct in GREEN_ROWS
            apply_subtotal(ws, row, max_col, green=is_green)
            ws.cell(row=row, column=2).font = F_BOLD
        else:
            if gl:
                ws.cell(row=row, column=1).font = F_MUTED
            if acct:
                ws.cell(row=row, column=2).font = F_BODY
            fill = P_ALT if row % 2 == 0 else P_WHITE
            for c in range(1, max_col + 1):
                ws.cell(row=row, column=c).fill = fill


def hide_beyond(ws, last_row, last_col, max_row=200, max_col=52):
    """Hide all rows and columns beyond the data boundary."""
    for col_idx in range(last_col + 1, max_col + 1):
        cl = get_column_letter(col_idx)
        ws.column_dimensions[cl].width = 0
        ws.column_dimensions[cl].hidden = True
    for col_idx in range(1, last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].hidden = False
    for row_idx in range(last_row + 1, max_row + 1):
        ws.row_dimensions[row_idx].hidden = True
    for row_idx in range(1, last_row + 1):
        ws.row_dimensions[row_idx].hidden = False


def set_col_widths(ws, widths):
    """Set column widths from a dict of {letter: width}."""
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


def input_cell(ws, row, col, value=None, nf=None):
    """Write a cell styled as an editable input (yellow bg, blue bold font)."""
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font = F_INPUT
    cell.fill = P_INPUT
    if nf:
        cell.number_format = nf
    return cell
